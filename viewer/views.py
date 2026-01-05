from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils.translation import gettext_lazy as _
from django.conf import settings
from qr_reader_django import crud
import json
from viewer.models import ScanEvent, Vacation
from qr_reader_django.audit import log_action, get_client_ip
from django.core.paginator import Paginator
from django.db.models import Q, Max
from django.views.decorators.csrf import csrf_exempt
import datetime
from django.utils.formats import date_format

# ============= PUBLIC VIEWS =============

def landing_page(request):
    """Landing page with links to company and user login"""
    # Redirect to dashboard if already logged in
    if 'company_id' in request.session and request.session.get('user_type') == 'company':
        return redirect('company_dashboard')
    elif 'user_id' in request.session and request.session.get('user_type') == 'user':
        return redirect('user_dashboard')
    
    return render(request, 'landing.html')


# ============= COMPANY AUTH VIEWS =============

def company_register(request):
    """Company registration page"""
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        auto_lunch_breaks = request.POST.get('auto_lunch_breaks') == 'on'

        if not all([name, email, password, confirm_password]):
            messages.error(request, _('All fields are required'))
            return render(request, 'company_register.html')

        if password != confirm_password:
            messages.error(request, _('Passwords do not match'))
            return render(request, 'company_register.html')

        company, error = crud.create_company(name, email, password, auto_lunch_breaks=auto_lunch_breaks, ip_address=get_client_ip(request))
        if error:
            messages.error(request, error)
            return render(request, 'company_register.html')

        messages.success(request, _('Company registered successfully! Please login.'))
        return redirect('company_login')

    return render(request, 'company_register.html')


def company_login(request):
    """Company login page"""
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        company = crud.get_company_by_email(email)
        if company and company.check_password(password):
            request.session['company_id'] = company.id
            request.session['user_type'] = 'company'
            
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='login',
                message=f'Company "{company.name}" logged in',
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, _('Welcome back, {}!').format(company.name))
            return redirect('company_dashboard')
        else:
            messages.error(request, _('Invalid credentials'))

    return render(request, 'company_login.html')


def company_logout(request):
    """Company logout"""
    if 'company_id' in request.session:
        company = crud.get_company_by_id(request.session['company_id'])
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='logout',
                message=f'Company "{company.name}" logged out',
                ip_address=get_client_ip(request)
            )
    
    request.session.flush()
    messages.success(request, _('Logged out successfully'))
    return redirect('landing_page')


def company_dashboard(request):
    """Company dashboard - manage QR codes, users, and absences"""
    # Allow both company owners and managers
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Please login as a company or manager'))
        return redirect('company_login')
    
    # Get company and current user
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
        current_user = None
    else:  # is_manager
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    # Determine permissions first
    can_edit_qr_codes = is_company or (current_user and current_user.can_edit_qr_codes)
    can_edit_employees = is_company or (current_user and current_user.can_edit_employees)
    can_edit_absences = is_company or (current_user and current_user.can_edit_absences)
    
    # Get active tab from query params
    requested_tab = request.GET.get('tab', '')
    
    # Validate tab permissions and set default if needed
    if requested_tab:
        # Check if user has permission for requested tab
        if requested_tab == 'qr-codes' and not can_edit_qr_codes:
            requested_tab = ''
        elif requested_tab == 'users' and not can_edit_employees:
            requested_tab = ''
        elif requested_tab == 'absences' and not can_edit_absences:
            requested_tab = ''
    
    # Set default tab if not specified or not permitted
    if not requested_tab:
        if can_edit_qr_codes:
            active_tab = 'qr-codes'
        elif can_edit_employees:
            active_tab = 'users'
        elif can_edit_absences:
            active_tab = 'absences'
        else:
            active_tab = 'qr-codes'  # fallback
    else:
        active_tab = requested_tab
    
    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    vacation_type = request.GET.get('vacation_type', '')
    user_filter = request.GET.get('user', '')
    absence_employee_name = request.GET.get('absence_employee_name', '')
    qr_name = request.GET.get('qr_name', '')
    employee_name = request.GET.get('employee_name', '')
    work_status = request.GET.get('work_status', '')
    items_per_page = request.GET.get('items_per_page', '25')
    sort = request.GET.get('sort', '')
    page_number = request.GET.get('page', 1)
    
    # Set default sorting based on active tab if not specified
    if not sort:
        if active_tab == 'qr-codes':
            sort = 'name'
        elif active_tab == 'users':
            sort = 'name'
        elif active_tab == 'absences':
            sort = '-date_from'
    
    # Validate items per page
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100]:
            items_per_page = 25
    except (ValueError, TypeError):
        items_per_page = 25
    
    # Get all QR codes for datalist (unfiltered)
    all_qr_codes = crud.get_company_qr_codes(company)
    
    # Get QR codes with filtering
    qr_codes = crud.get_company_qr_codes(company)
    if qr_name:
        qr_codes = qr_codes.filter(name__icontains=qr_name)
    
    # Apply sorting for QR codes (default: ASC on name)
    if active_tab == 'qr-codes':
        if sort:
            if sort == 'name':
                qr_codes = qr_codes.order_by('name')
            elif sort == '-name':
                qr_codes = qr_codes.order_by('-name')
            elif sort == 'created_at':
                qr_codes = qr_codes.order_by('created_at')
            elif sort == '-created_at':
                qr_codes = qr_codes.order_by('-created_at')
        else:
            # Default sort: ASC on name
            qr_codes = qr_codes.order_by('name')
    
    # Get all users for datalist (unfiltered)
    all_users = crud.get_company_users(company)
    
    # Get users with filtering
    users = crud.get_company_users(company)
    if employee_name:
        users = users.filter(name__icontains=employee_name)
    
    # Calculate arrivals and departures for each QR code and annotate scan count
    qr_codes_list = list(qr_codes)
    for qr in qr_codes_list:
        qr.arrivals_count = qr.scans.filter(scan_type='arrival').count()
        qr.departures_count = qr.scans.filter(scan_type='departure').count()
        qr.total_scans = qr.scans.count()
    
    # Apply sorting for QR codes (on scan count - requires list)
    if active_tab == 'qr-codes' and sort:
        if sort == 'scans':
            qr_codes_list.sort(key=lambda x: x.total_scans)
        elif sort == '-scans':
            qr_codes_list.sort(key=lambda x: x.total_scans, reverse=True)
    
    # Paginate QR codes
    if active_tab == 'qr-codes':
        qr_paginator = Paginator(qr_codes_list, items_per_page)
        qr_codes_page = qr_paginator.get_page(page_number)
    else:
        qr_codes_page = None
    
    # Calculate total scans for each user (only from active QR codes)
    # Store users as list to allow filtering by work status
    users_list = list(users)
    for user in users_list:
        user.total_scans = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__is_active=True
        ).count()
        
        # Get last scan to determine if user is at work
        last_scan = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__is_active=True
        ).order_by('-timestamp').first()
        
        if last_scan:
            user.is_at_work = last_scan.scan_type == 'arrival'
            user.work_location = last_scan.address if last_scan.address else f"{last_scan.latitude}, {last_scan.longitude}"
        else:
            user.is_at_work = False
            user.work_location = None
    
    # Apply work status filter
    if work_status:
        if work_status == 'at_work':
            users_list = [u for u in users_list if u.is_at_work]
        elif work_status == 'not_at_work':
            users_list = [u for u in users_list if not u.is_at_work]
    
    # Apply sorting for users (default: ASC on name)
    if active_tab == 'users':
        if sort:
            if sort == 'name':
                users_list.sort(key=lambda x: x.name.lower())
            elif sort == '-name':
                users_list.sort(key=lambda x: x.name.lower(), reverse=True)
            elif sort == 'scans':
                users_list.sort(key=lambda x: x.total_scans)
            elif sort == '-scans':
                users_list.sort(key=lambda x: x.total_scans, reverse=True)
            elif sort == 'at_work':
                users_list.sort(key=lambda x: x.is_at_work)
            elif sort == '-at_work':
                users_list.sort(key=lambda x: x.is_at_work, reverse=True)
        else:
            # Default sort: ASC on name
            users_list.sort(key=lambda x: x.name.lower())
    
    # Paginate users
    if active_tab == 'users':
        users_paginator = Paginator(users_list, items_per_page)
        users_page = users_paginator.get_page(page_number)
    else:
        users_page = None
    
    # Get absences for the company
    absences = Vacation.objects.filter(
        user__company=company,
        user__is_active=True
    ).select_related('user')
    
    # Apply filters to absences
    if date_from and date_to:
        try:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            # Filter absences that overlap with the selected date range
            # Absence overlaps if it starts before the end of range AND ends after the start of range
            absences = absences.filter(
                date_from__lte=date_to_obj,
                date_to__gte=date_from_obj
            )
        except ValueError:
            pass
    
    if vacation_type and vacation_type != 'all':
        absences = absences.filter(type=vacation_type)
    
    if user_filter:
        absences = absences.filter(user__id=user_filter)
    
    if absence_employee_name:
        absences = absences.filter(user__name__icontains=absence_employee_name)
    
    # Apply sorting for absences (default: DESC on date_from)
    if active_tab == 'absences':
        if sort:
            if sort == 'name':
                absences = absences.order_by('user__name')
            elif sort == '-name':
                absences = absences.order_by('-user__name')
            elif sort == 'date_from':
                absences = absences.order_by('date_from')
            elif sort == '-date_from':
                absences = absences.order_by('-date_from')
            elif sort == 'date_to':
                absences = absences.order_by('date_to')
            elif sort == '-date_to':
                absences = absences.order_by('-date_to')
            elif sort == 'type':
                absences = absences.order_by('type')
            elif sort == '-type':
                absences = absences.order_by('-type')
        else:
            # Default sort: DESC on date_from
            absences = absences.order_by('-date_from')
    
    # Paginate absences
    if active_tab == 'absences':
        absences_paginator = Paginator(absences, items_per_page)
        absences_page = absences_paginator.get_page(page_number)
    else:
        absences_page = None

    users_json = json.dumps([{'id': u.id, 'name': u.name} for u in all_users])

    context = {
        'company': company,
        'current_user': current_user,
        'is_company': is_company,
        'is_manager': is_manager,
        'can_edit_qr_codes': can_edit_qr_codes,
        'can_edit_employees': can_edit_employees,
        'can_edit_absences': can_edit_absences,
        'all_qr_codes': all_qr_codes,  # All QR codes for datalist
        'qr_codes': qr_codes_list if active_tab == 'qr-codes' else [],
        'users': all_users,  # All users for datalist
        'users_json': users_json,
        'users_list': users_list if active_tab == 'users' else [],
        'absences': absences if active_tab == 'absences' else [],
        'qr_codes_page': qr_codes_page,
        'users_page': users_page,
        'absences_page': absences_page,
        'active_tab': active_tab,
        'current_filters': {
            'date_from': date_from,
            'date_to': date_to,
            'vacation_type': vacation_type,
            'user': user_filter,
            'absence_employee_name': absence_employee_name,
            'qr_name': qr_name,
            'employee_name': employee_name,
            'work_status': work_status,
            'items_per_page': str(items_per_page),
            'sort': sort,
        },
        'qr_codes_count': len(qr_codes_list),
        'users_count': len(users_list),
        'absences_count': absences.count(),
    }
    return render(request, 'company_dashboard.html', context)


# ============= USER AUTH VIEWS =============

def user_login(request):
    """User login page"""
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        print(f"USER LOGIN ATTEMPT - Email: {email}, DEBUG: {settings.DEBUG}")

        user = crud.get_user_by_email(email)
        print(f"User found: {user}, User active: {user.is_active if user else 'N/A'}")
        
        if user and user.check_password(password):
            print(f"Password check passed for user: {user.name}")
            request.session['user_id'] = user.id
            request.session['user_type'] = 'user'
            
            log_action(
                actor_type='user',
                actor_email=user.email,
                actor_name=user.name,
                action='login',
                message=f'User "{user.name}" logged in',
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, _('Welcome back, {}!').format(user.name))
            return redirect('user_dashboard')
        else:
            print(f"Login failed - User: {user}, Password check: {user.check_password(password) if user else 'N/A'}")
            messages.error(request, _('Invalid credentials'))

    return render(request, 'user_login.html')


def user_logout(request):
    """User logout"""
    if 'user_id' in request.session:
        user = crud.get_user_by_id(request.session['user_id'])
        if user:
            log_action(
                actor_type='user',
                actor_email=user.email,
                actor_name=user.name,
                action='logout',
                message=f'User "{user.name}" logged out',
                ip_address=get_client_ip(request)
            )
    
    request.session.flush()
    messages.success(request, _('Logged out successfully'))
    return redirect('landing_page')


def user_dashboard(request):
    if 'user_id' not in request.session or request.session.get('user_type') != 'user':
        messages.error(request, _('Please login as a user'))
        return redirect('user_login')

    user = crud.get_user_by_id(request.session['user_id'])
    if not user:
        messages.error(request, _('User not found'))
        return redirect('user_login')
    
    # Get active tab from query parameter, default to 'scans'
    active_tab = request.GET.get('tab', 'scans')
    
    # Get unique QR codes for filter dropdown
    qr_codes = crud.get_company_qr_codes(user.company)
    qr_code_names = [qr.name for qr in qr_codes]
    
    # Shared filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort', '-timestamp' if active_tab == 'scans' else '-date_from')
    page_number = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 20)
    
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 50, 100]:
            per_page = 20
    except:
        per_page = 20
    
    # ===== SCANS TAB =====
    scans = ScanEvent.objects.filter(
        qr_code__company=user.company,
        scanned_by=user,
        qr_code__is_active=True
    ).select_related('qr_code', 'scanned_by')
    
    # Scans-specific filters
    qr_code_filter = request.GET.get('qr_code', '')
    scan_type_filter = request.GET.get('scan_type', '')
    
    if qr_code_filter:
        scans = scans.filter(qr_code__name__icontains=qr_code_filter)
    
    if scan_type_filter:
        scans = scans.filter(scan_type=scan_type_filter)
    
    if date_from:
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d')
        scans = scans.filter(timestamp__date__gte=date_from_obj.date())
    
    if date_to:
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d')
        scans = scans.filter(timestamp__date__lte=date_to_obj.date())
    
    # Sorting for scans
    valid_scan_sort_fields = ['timestamp', '-timestamp', 'qr_code__name', '-qr_code__name', 
                         'scan_type', '-scan_type']
    
    if active_tab == 'scans' and sort_by in valid_scan_sort_fields:
        scans = scans.order_by(sort_by)
    else:
        scans = scans.order_by('-timestamp')
    
    # Pagination for scans
    scans_paginator = Paginator(scans, per_page)
    page_obj = scans_paginator.get_page(page_number)
    scans_count = scans.count()
    
    # ===== ABSENCES TAB =====
    absences = Vacation.objects.filter(user=user).order_by('-date_from')
    
    # Absences-specific filters
    vacation_type_filter = request.GET.get('vacation_type', '')
    
    if vacation_type_filter:
        absences = absences.filter(type=vacation_type_filter)
    
    if date_from and date_to:
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
        absences = absences.filter(
            date_from__lte=date_to_obj,
            date_to__gte=date_from_obj
        )
    
    # Sorting for absences
    valid_absence_sort_fields = ['date_from', '-date_from', 'date_to', '-date_to', 
                                  'created_at', '-created_at']
    
    if active_tab == 'absences' and sort_by in valid_absence_sort_fields:
        absences = absences.order_by(sort_by)
    else:
        absences = absences.order_by('-date_from')
    
    # Pagination for absences
    absences_paginator = Paginator(absences, per_page)
    absences_page = absences_paginator.get_page(page_number)
    absences_count = absences.count()
    
    # Check if any filters are active
    has_active_filters = any([qr_code_filter, scan_type_filter, vacation_type_filter, date_from, date_to])

    context = {
        'user': user,
        'page_obj': page_obj,
        'absences_page': absences_page,
        'scans_count': scans_count,
        'absences_count': absences_count,
        'active_tab': active_tab,
        'qr_codes': qr_codes,
        'has_active_filters': has_active_filters,
        'datalist_items': qr_code_names,
        'current_filters': {
            'qr_code': qr_code_filter,
            'scan_type': scan_type_filter,
            'vacation_type': vacation_type_filter,
            'date_from': date_from,
            'date_to': date_to,
            'sort': sort_by,
            'per_page': per_page,
        }
    }
    return render(request, 'user_dashboard.html', context)


def user_scan_qr(request):
    """User QR scanner page"""
    if 'user_id' not in request.session or request.session.get('user_type') != 'user':
        messages.error(request, _('Please login as a user'))
        return redirect('user_login')

    user = crud.get_user_by_id(request.session['user_id'])
    if not user:
        messages.error(request, _('User not found'))
        return redirect('user_login')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            uuid = data.get('uuid')
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            scan_type = data.get('scan_type', 'arrival')
            
            # Check if QR code exists and belongs to user's company
            qr_code = crud.get_qr_code_by_uuid(uuid)
            
            if not qr_code:
                return JsonResponse({
                    'status': 'error',
                    'message': str(_('QR code not found or inactive'))
                }, status=404)
            
            if qr_code.company != user.company:
                return JsonResponse({
                    'status': 'error',
                    'message': str(_('This QR code does not belong to your company'))
                }, status=403)
            
            # Record the scan with audit logging
            scan, address = crud.create_scan_event(
                qr_code=qr_code,
                scanned_by=user,
                latitude=latitude,
                longitude=longitude,
                scan_type=scan_type,
                device_info=request.META.get('HTTP_USER_AGENT', ''),
                actor_type='user',
                actor_email=user.email,
                actor_name=user.name,
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('Scan recorded successfully!')),
                'data': {
                    'qr_name': qr_code.name,
                    'qr_location': qr_code.location,
                    'scan_timestamp': scan.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    'scan_latitude': latitude,
                    'scan_longitude': longitude,
                    'scan_address': address or str(_('Address not available'))
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
    
    # Determine which buttons should be enabled based on today's scans
    from datetime import date
    from django.utils import timezone
    
    today = timezone.now().date()
    today_scans = ScanEvent.objects.filter(
        scanned_by=user,
        timestamp__date=today
    ).order_by('timestamp')
    
    # Default: only arrival enabled
    enabled_buttons = ['arrival']
    
    if today_scans.exists():
        last_scan = today_scans.last()
        
        if last_scan.scan_type == 'arrival':
            # After arrival: can depart or start lunch break
            enabled_buttons = ['departure', 'lunch_break_start']
        elif last_scan.scan_type == 'lunch_break_start':
            # During lunch break: can only end lunch break
            enabled_buttons = ['lunch_break_end']
        elif last_scan.scan_type == 'lunch_break_end':
            # After lunch break: can depart or start another lunch break
            enabled_buttons = ['departure', 'lunch_break_start']
        elif last_scan.scan_type == 'departure':
            # After departure: can arrive again (new shift)
            enabled_buttons = ['arrival']
    
    context = {
        'user': user,
        'enabled_buttons': enabled_buttons,
    }
    return render(request, 'user_scan_qr.html', context)


# ============= COMPANY ACTIONS =============

def create_qr_code(request):
    """Create a new QR code (company or manager with permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get company
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
            else:
                user = crud.get_user_by_id(request.session['user_id'])
                if not user or not user.is_manager or not user.can_edit_qr_codes:
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
                company = user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                actor_type = 'user'
                actor_email = user.email
                actor_name = user.name
            
            qr_code, error = crud.create_qr_code(
                company=company,
                name=data.get('name'),
                location=data.get('location'),
                additional_info=data.get('additional_info', ''),
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if error:
                return JsonResponse({'status': 'error', 'message': error}, status=400)
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('QR code created successfully')),
                'qr_code_id': qr_code.id,
                'uuid': qr_code.uuid
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def delete_qr_code(request, qr_id):
    """Delete/deactivate a QR code (company or manager with permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        user = crud.get_user_by_id(request.session['user_id'])
        if not user or not user.is_manager or not user.can_edit_qr_codes:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    # Extract actor info for audit logging
    if is_company:
        actor_type = 'company'
        actor_email = company.email
        actor_name = company.name
    else:
        actor_type = 'user'
        actor_email = user.email
        actor_name = user.name
    
    success, error = crud.deactivate_qr_code(
        qr_id, 
        company,
        actor_type=actor_type,
        actor_email=actor_email,
        actor_name=actor_name,
        ip_address=get_client_ip(request)
    )
    if success:
        messages.success(request, _('QR code deactivated successfully'))
    else:
        messages.error(request, error or _('Failed to deactivate QR code'))
    
    return redirect('company_dashboard')


def create_user(request):
    """Register a new user under the company (company or manager with permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get company
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
            else:
                user = crud.get_user_by_id(request.session['user_id'])
                if not user or not user.is_manager or not user.can_edit_employees:
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
                company = user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            # Get actor info
            if is_company:
                actor_type, actor_email, actor_name = 'company', company.email, company.name
            else:
                actor_type, actor_email, actor_name = 'user', user.email, user.name
            
            user, error = crud.create_user(
                company=company,
                name=data.get('name'),
                email=data.get('email'),
                password=data.get('password'),
                basic_work_hours=data.get('basic_work_hours', 160),
                holidays_per_year=data.get('holidays_per_year', 20),
                is_manager=data.get('is_manager', False),
                can_edit_employees=data.get('can_edit_employees', False),
                can_edit_qr_codes=data.get('can_edit_qr_codes', False),
                can_edit_absences=data.get('can_edit_absences', False),
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if error:
                return JsonResponse({'status': 'error', 'message': error}, status=400)
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('User created successfully')),
                'user_id': user.id
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def edit_user(request, user_id):
    """Edit user details (company or manager with permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get company
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
            else:
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
                company = current_user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            user = crud.get_user_by_id(user_id)
            if not user or user.company != company:
                return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
            
            data = json.loads(request.body)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                manager = crud.get_user_by_id(request.session['user_id'])
                actor_type = 'user'
                actor_email = manager.email
                actor_name = manager.name
            
            user, error = crud.update_user(
                user_id=user_id,
                company=company,
                name=data.get('name'),
                email=data.get('email'),
                password=data.get('password'),
                basic_work_hours=data.get('basic_work_hours'),
                holidays_per_year=data.get('holidays_per_year'),
                is_active=data.get('is_active'),
                is_manager=data.get('is_manager'),
                can_edit_employees=data.get('can_edit_employees'),
                can_edit_qr_codes=data.get('can_edit_qr_codes'),
                can_edit_absences=data.get('can_edit_absences'),
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if error:
                return JsonResponse({'status': 'error', 'message': error}, status=400)
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('User updated successfully'))
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def delete_user(request, user_id):
    """Delete user (company or manager with permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            # Get company
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
            else:
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
                company = current_user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            user = crud.get_user_by_id(user_id)
            if not user or user.company != company:
                return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                manager = crud.get_user_by_id(request.session['user_id'])
                actor_type = 'user'
                actor_email = manager.email
                actor_name = manager.name
            
            success, error = crud.delete_user(
                user_id, 
                company,
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if success:
                return JsonResponse({
                    'status': 'success',
                    'message': str(_('User deleted successfully'))
                })
            else:
                return JsonResponse({'status': 'error', 'message': error}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def view_qr_scans(request, qr_id):
    """View all scans for a specific QR code with filtering and pagination"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        user = crud.get_user_by_id(request.session['user_id'])
        if not user or not user.is_manager or not user.can_edit_qr_codes:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    qr_code = crud.get_qr_code_by_id(qr_id, company)
    if not qr_code:
        messages.error(request, _('QR code not found'))
        return redirect('company_dashboard')
    
    scans = ScanEvent.objects.filter(qr_code=qr_code).select_related('qr_code', 'scanned_by')
    
    # Filtering
    user_filter = request.GET.get('user', '')
    scan_type_filter = request.GET.get('scan_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if user_filter:
        scans = scans.filter(scanned_by__name__icontains=user_filter)
    
    if scan_type_filter:
        scans = scans.filter(scan_type=scan_type_filter)
    
    if date_from:
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d')
        scans = scans.filter(timestamp__date__gte=date_from_obj.date())
    
    if date_to:
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d')
        scans = scans.filter(timestamp__date__lte=date_to_obj.date())
    
    # Sorting - always default to DESC by timestamp if not specified
    sort_by = request.GET.get('sort', '-timestamp')
    valid_sort_fields = ['timestamp', '-timestamp', 'scanned_by__name', '-scanned_by__name', 
                         'scan_type', '-scan_type']
    
    if sort_by in valid_sort_fields:
        scans = scans.order_by(sort_by)
    else:
        scans = scans.order_by('-timestamp')
    
    # Pagination
    page_number = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 20)
    
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 50, 100]:
            per_page = 20
    except:
        per_page = 20
    
    paginator = Paginator(scans, per_page)
    page_obj = paginator.get_page(page_number)
    
    # Check if any filters are active
    has_active_filters = any([user_filter, scan_type_filter, date_from, date_to])
    
    # Get all users for datalist
    users = crud.get_company_users(company)
    user_names = [u.name for u in users]

    context = {
        'company': company,
        'qr_code': qr_code,
        'page_obj': page_obj,
        'has_active_filters': has_active_filters,
        'datalist_items': user_names,
        'current_filters': {
            'user': user_filter,
            'scan_type': scan_type_filter,
            'date_from': date_from,
            'date_to': date_to,
            'sort': sort_by,
            'per_page': per_page,
        }
    }
    return render(request, 'qr_scans.html', context)


def view_user_details(request, user_id):
    """View detailed information about a specific user with filtering and pagination"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    user = crud.get_user_by_id(user_id)
    if not user or user.company != company:
        messages.error(request, _('User not found or inactive'))
        return redirect('company_dashboard')
    
    # Determine active tab
    active_tab = request.GET.get('tab', 'scans')
    
    # Filtering
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Pagination
    page_number = request.GET.get('page', 1)
    per_page = request.GET.get('per_page', 20)
    
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 50, 100]:
            per_page = 20
    except:
        per_page = 20
    
    if active_tab == 'vacations':
        # Get all vacations by this user
        vacations = Vacation.objects.filter(user=user, is_active=True)
        
        # Additional filters for vacations
        vacation_type_filter = request.GET.get('vacation_type', '')
        
        if vacation_type_filter:
            vacations = vacations.filter(type=vacation_type_filter)
        
        # Date filter: show vacations that overlap with the date range
        if date_from and date_to:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            vacations = vacations.filter(
                date_from__lte=date_to_obj,
                date_to__gte=date_from_obj
            )
        elif date_from:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            vacations = vacations.filter(date_to__gte=date_from_obj)
        elif date_to:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            vacations = vacations.filter(date_from__lte=date_to_obj)
        
        # Sorting - default to DESC by date_from
        sort_by = request.GET.get('sort', '-date_from')
        valid_sort_fields = ['date_from', '-date_from', 'date_to', '-date_to', 
                             'created_at', '-created_at']
        
        if sort_by in valid_sort_fields:
            vacations = vacations.order_by(sort_by)
        else:
            vacations = vacations.order_by('-date_from')
        
        # Pagination
        paginator = Paginator(vacations, per_page)
        vacations_page = paginator.get_page(page_number)
        
        # Check if any filters are active
        has_active_filters = any([date_from, date_to, vacation_type_filter])
        
        # Get counts
        scans_count = ScanEvent.objects.filter(scanned_by=user).count()
        vacations_count = Vacation.objects.filter(user=user).count()
        
        context = {
            'company': company,
            'user': user,
            'vacations_page': vacations_page,
            'page_obj': vacations_page,  # For paginator template
            'has_active_filters': has_active_filters,
            'active_tab': active_tab,
            'scans_count': scans_count,
            'vacations_count': vacations_count,
            'current_filters': {
                'date_from': date_from,
                'date_to': date_to,
                'vacation_type': vacation_type_filter,
                'sort': sort_by,
                'per_page': per_page,
            }
        }
    else:
        # Get all scans by this user
        scans = ScanEvent.objects.filter(scanned_by=user).select_related('qr_code', 'scanned_by')
        
        # Additional filters for scans
        qr_code_filter = request.GET.get('qr_code', '')
        scan_type_filter = request.GET.get('scan_type', '')
        
        if qr_code_filter:
            scans = scans.filter(qr_code__name__icontains=qr_code_filter)
        
        if scan_type_filter:
            scans = scans.filter(scan_type=scan_type_filter)
        
        if date_from:
            date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d')
            scans = scans.filter(timestamp__date__gte=date_from_obj.date())
        
        if date_to:
            date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d')
            scans = scans.filter(timestamp__date__lte=date_to_obj.date())
        
        # Sorting - always default to DESC by timestamp
        sort_by = request.GET.get('sort', '-timestamp')
        valid_sort_fields = ['timestamp', '-timestamp', 'qr_code__name', '-qr_code__name', 
                             'scan_type', '-scan_type']
        
        if sort_by in valid_sort_fields:
            scans = scans.order_by(sort_by)
        else:
            scans = scans.order_by('-timestamp')
        
        # Pagination
        paginator = Paginator(scans, per_page)
        page_obj = paginator.get_page(page_number)
        
        # Check if any filters are active
        has_active_filters = any([qr_code_filter, scan_type_filter, date_from, date_to])
        
        # Get all QR codes for datalist
        qr_codes = crud.get_company_qr_codes(company)
        qr_code_names = [qr.name for qr in qr_codes]
        
        # Get counts
        scans_count = ScanEvent.objects.filter(scanned_by=user).count()
        vacations_count = Vacation.objects.filter(user=user).count()

        context = {
            'company': company,
            'user': user,
            'page_obj': page_obj,
            'vacations_page': page_obj,  # Empty for scans tab
            'has_active_filters': has_active_filters,
            'datalist_items': qr_code_names,
            'active_tab': active_tab,
            'scans_count': scans_count,
            'vacations_count': vacations_count,
            'current_filters': {
                'qr_code': qr_code_filter,
                'scan_type': scan_type_filter,
                'date_from': date_from,
                'date_to': date_to,
                'sort': sort_by,
                'per_page': per_page,
            }
        }
    return render(request, 'company_user_details.html', context)

def create_vacation(request):
    """Create a new vacation (company, manager, or user for themselves)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_user = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_user):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get company and current user
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
                current_user = None
            else:  # is_user (regular user or manager)
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user:
                    return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
                company = current_user.company
                
                # Check if manager has permission to edit absences
                if current_user.is_manager and not current_user.can_edit_absences:
                    # Manager without permission can only create for themselves
                    if data.get('user_id') != current_user.id:
                        return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            # Determine target user
            target_user_id = data.get('user_id')
            
            # Regular users can only create vacations for themselves
            if is_user and not is_company:
                if current_user.is_manager and current_user.can_edit_absences:
                    # Manager with permission can create for anyone in company
                    user = crud.get_user_by_id(target_user_id)
                else:
                    # Regular user or manager without permission can only create for themselves
                    user = current_user
            else:
                # Company can create for anyone
                user = crud.get_user_by_id(target_user_id)
            
            if not user or user.company != company:
                return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
            
            # Set approved=True if created by company or manager with permission, False otherwise
            approved = is_company or (current_user and current_user.is_manager and current_user.can_edit_absences)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                actor_type = 'user'
                actor_email = current_user.email
                actor_name = current_user.name
            
            vacation, error = crud.create_vacation(
                user=user,
                date_from=data.get('date_from'),
                date_to=data.get('date_to'),
                vacation_type=data.get('type', 'vacation'),
                approved=approved,
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if error:
                return JsonResponse({'status': 'error', 'message': error}, status=400)
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('Vacation created successfully')),
                'vacation_id': vacation.id
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def edit_vacation(request, vacation_id):
    """Edit vacation details (company, manager with permission, or user editing their own)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_user = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_user):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get company and current user
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
                current_user = None
            else:  # is_user
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user:
                    return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
                company = current_user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            # Get the vacation to check ownership
            try:
                vacation = Vacation.objects.get(id=vacation_id, user__company=company)
            except Vacation.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': str(_('Vacation not found'))}, status=404)
            
            # Check permissions
            if is_user and not is_company:
                # Regular user can only edit their own vacations
                if current_user.is_manager and current_user.can_edit_absences:
                    # Manager with permission can edit anyone's vacation
                    pass
                elif vacation.user != current_user:
                    # Regular user or manager without permission can only edit their own
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                actor_type = 'user'
                actor_email = current_user.email
                actor_name = current_user.name
            
            vacation, error = crud.update_vacation(
                vacation_id=vacation_id,
                company=company,
                user_id=data.get('user_id'),
                date_from=data.get('date_from'),
                date_to=data.get('date_to'),
                vacation_type=data.get('type'),
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if error:
                return JsonResponse({'status': 'error', 'message': error}, status=400)
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('Vacation updated successfully'))
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def delete_vacation(request, vacation_id):
    """Delete vacation (company, manager with permission, or user deleting their own)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_user = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_user):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            # Get company and current user
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
                current_user = None
            else:  # is_user
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user:
                    return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
                company = current_user.company
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            # Get the vacation to check ownership
            try:
                vacation = Vacation.objects.get(id=vacation_id, user__company=company)
            except Vacation.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': str(_('Vacation not found'))}, status=404)
            
            # Check permissions
            if is_user and not is_company:
                # Regular user can only delete their own vacations
                if current_user.is_manager and current_user.can_edit_absences:
                    # Manager with permission can delete anyone's vacation
                    pass
                elif vacation.user != current_user:
                    # Regular user or manager without permission can only delete their own
                    return JsonResponse({'status': 'error', 'message': str(_('Access denied'))}, status=403)
            
            # Extract actor info for audit logging
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                actor_type = 'user'
                actor_email = current_user.email
                actor_name = current_user.name
            
            success, error = crud.delete_vacation(
                vacation_id, 
                company,
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                ip_address=get_client_ip(request)
            )
            
            if success:
                return JsonResponse({
                    'status': 'success',
                    'message': str(_('Vacation deleted successfully'))
                })
            else:
                return JsonResponse({'status': 'error', 'message': error}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def approve_vacation(request, vacation_id):
    """Approve vacation (company or manager with can_edit_absences permission)"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_user = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_user):
        return JsonResponse({'status': 'error', 'message': str(_('Unauthorized'))}, status=403)

    if request.method == 'POST':
        try:
            # Get company and current user
            if is_company:
                company = crud.get_company_by_id(request.session['company_id'])
                can_approve = True
            else:  # is_user
                current_user = crud.get_user_by_id(request.session['user_id'])
                if not current_user:
                    return JsonResponse({'status': 'error', 'message': str(_('User not found'))}, status=404)
                company = current_user.company
                # Check if user is manager with can_edit_absences permission
                can_approve = current_user.is_manager and current_user.can_edit_absences
            
            if not company:
                return JsonResponse({'status': 'error', 'message': str(_('Company not found'))}, status=404)
            
            if not can_approve:
                return JsonResponse({'status': 'error', 'message': str(_('Access denied. Only company or managers with edit absences permission can approve.'))}, status=403)
            
            # Get the vacation
            try:
                vacation = Vacation.objects.get(id=vacation_id, user__company=company)
            except Vacation.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': str(_('Vacation not found'))}, status=404)
            
            # Set approved to True
            vacation.approved = True
            vacation.save()
            
            # Log approval action
            if is_company:
                actor_type = 'company'
                actor_email = company.email
                actor_name = company.name
            else:
                actor_type = 'user'
                actor_email = current_user.email
                actor_name = current_user.name
            
            log_action(
                actor_type=actor_type,
                actor_email=actor_email,
                actor_name=actor_name,
                action='approve',
                message=f'Approved vacation for {vacation.user.name} ({vacation.date_from} to {vacation.date_to})',
                ip_address=get_client_ip(request)
            )
            
            return JsonResponse({
                'status': 'success',
                'message': str(_('Vacation approved successfully'))
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': str(_('Invalid method'))}, status=405)


def generate_attendance_pdf(request, user_id):
    """Generate PDF attendance report for a user"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime, timedelta
    from collections import defaultdict
    from calendar import monthrange
    import holidays
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    user = crud.get_user_by_id(user_id)
    if not user or user.company != company:
        messages.error(request, _('User not found'))
        return redirect('company_dashboard')
    
    # Parse date range - accept both date_range and date_from/date_to parameters
    date_range = request.GET.get('date_range', '')
    date_from_param = request.GET.get('date_from', '')
    date_to_param = request.GET.get('date_to', '')
    
    try:
        if date_from_param and date_to_param:
            # Use date_from and date_to parameters (format: YYYY-MM-DD)
            date_from = datetime.strptime(date_from_param, '%Y-%m-%d')
            date_to = datetime.strptime(date_to_param, '%Y-%m-%d')
        elif date_range and ' - ' in date_range:
            # Use date_range parameter (format: DD.MM.YYYY - DD.MM.YYYY)
            date_from_str, date_to_str = date_range.split(' - ')
            date_from = datetime.strptime(date_from_str.strip(), '%d.%m.%Y')
            date_to = datetime.strptime(date_to_str.strip(), '%d.%m.%Y')
        else:
            messages.error(request, _('Invalid date range'))
            return redirect('view_user_details', user_id=user_id)
    except:
        messages.error(request, _('Invalid date format'))
        return redirect('view_user_details', user_id=user_id)
    
    # Get scans in date range
    scans = ScanEvent.objects.filter(
        scanned_by=user,
        timestamp__date__gte=date_from.date(),
        timestamp__date__lte=date_to.date()
    ).select_related('qr_code').order_by('timestamp')
    
    vacations = Vacation.objects.filter(
        user=user,
        is_active=True,
        date_from__lte=date_to.date(),
        date_to__gte=date_from.date()
    ).order_by('date_from')
    
    # Create directory structure for PDF storage
    now = datetime.now()
    pdf_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Generate filename
    filename = f"attendance_{user.name.replace(' ', '_')}_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.pdf"
    filepath = os.path.join(pdf_dir, filename)
    
    # Register DejaVu fonts for Unicode support (Slovak characters)
    try:
        # Use fonts from project directory
        font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
        dejavu_path = os.path.join(font_dir, 'DejaVuSans.ttf')
        dejavu_bold_path = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
        
        if os.path.exists(dejavu_path) and os.path.exists(dejavu_bold_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
            font_name = 'DejaVuSans'
            font_name_bold = 'DejaVuSans-Bold'
        else:
            raise Exception("DejaVu fonts not found in project")
    except Exception as e:
        # Fallback to Helvetica if DejaVu is not available
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
    
    # Create PDF
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), 
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Unicode-compatible font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=20,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_name_bold,
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=11
    )
    
    # Title
    elements.append(Paragraph(f"{_('Attendance Report')} - {user.name}", title_style))
    elements.append(Paragraph(f"{_('Period')}: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}", normal_style))
    elements.append(Paragraph(f"{_('Company')}: {company.name}", normal_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Group scans by day
    daily_data = defaultdict(list)
    for scan in scans:
        day = scan.timestamp.date()
        daily_data[day].append(scan)
    
    # Create dictionary of vacation days with type
    vacation_days = {}
    for vacation in vacations:
        current = vacation.date_from
        while current <= vacation.date_to:
            vacation_days[current] = vacation.type if vacation.type else 'vacation'
            current += timedelta(days=1)
    
    # Helper function to calculate night hours (22:00-06:00)
    def calculate_night_hours(start_time, end_time):
        """Calculate hours worked between 22:00 and 06:00"""
        night_hours = 0
        current = start_time
        
        while current < end_time:
            # Check if current hour is night time (22:00-23:59 or 00:00-05:59)
            hour = current.hour
            if hour >= 22 or hour < 6:
                # Calculate minutes in this hour that count as night work
                next_hour = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
                segment_end = min(next_hour, end_time)
                segment_duration = (segment_end - current).total_seconds() / 3600
                night_hours += segment_duration
            
            # Move to next hour
            current = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            if current >= end_time:
                break
        
        return night_hours
    
    # Calculate statistics
    total_days = len(daily_data)
    total_work_hours = 0
    total_night_hours = 0
    days_with_issues = []
    total_vacation_days = 0
    
    # Daily attendance table
    elements.append(Paragraph(str(_('Daily Attendance')), heading_style))
    
    # Use Paragraph objects for header cells to enable wrapping
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=font_name_bold,
        fontSize=9,
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
        leading=11
    )
    
    table_data = [[
        Paragraph(str(_('Date')), header_style),
        Paragraph(str(_('Day')), header_style),
        Paragraph(str(_('Arrival')), header_style),
        Paragraph(str(_('Departure')), header_style),
        Paragraph(str(_('Hours')), header_style),
        Paragraph(str(_('Break')), header_style),
        Paragraph(str(_('Scanned QR')), header_style),
        Paragraph(str(_('Notes')), header_style)
    ]]
    
    # Style for data cells
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        leading=11
    )
    
    cell_style_centered = ParagraphStyle(
        'CellStyleCentered',
        parent=cell_style,
        alignment=TA_CENTER
    )
    
    # Get holidays based on language
    lang_code = request.LANGUAGE_CODE if hasattr(request, 'LANGUAGE_CODE') else 'sk'
    if lang_code == 'sk':
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    elif lang_code == 'en':
        country_holidays = holidays.UnitedStates(years=[date_from.year, date_to.year])
    elif lang_code == 'de':
        country_holidays = holidays.Germany(years=[date_from.year, date_to.year])
    elif lang_code == 'es':
        country_holidays = holidays.Spain(years=[date_from.year, date_to.year])
    else:
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    
    total_holiday_hours = 0
    
    current_date = date_from.date()
    while current_date <= date_to.date():
        day_scans = daily_data.get(current_date, [])
        # Use Django's date_format with 'l' format (day of the week)
        day_name = date_format(current_date, format='l')
        
        # Check if this day is a holiday
        is_holiday = current_date in country_holidays
        holiday_name = country_holidays.get(current_date, '') if is_holiday else ''
        
        # Check if this day is a vacation day
        vacation_type = vacation_days.get(current_date)
        is_vacation = vacation_type is not None
        if is_vacation:
            total_vacation_days += 1
        
        if is_vacation and not day_scans:
            # Vacation day with no scans - display type
            if vacation_type == 'sick_leave':
                vacation_style = ParagraphStyle(
                    'SickLeaveStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#f59e0b'),
                    fontName=font_name_bold
                )
                leave_label = f"🏥 {_('Sick Leave')}"
            elif vacation_type == 'doctor':
                vacation_style = ParagraphStyle(
                    'DoctorStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#8b5cf6'),
                    fontName=font_name_bold
                )
                leave_label = f"👨‍⚕️ {_('Doctor')}"
            else:
                vacation_style = ParagraphStyle(
                    'VacationStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#10b981'),
                    fontName=font_name_bold
                )
                leave_label = f"🏖 {_('Vacation')}"
            
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style),
                Paragraph(leave_label, vacation_style)
            ])
        elif not day_scans and not is_vacation:
            # No scans for this day
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('0:00', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style),
                Paragraph(str(_('No scans')), cell_style)
            ])
        else:
            # Day with scans (may or may not be vacation)
            # Find arrivals and departures
            arrivals = [s for s in day_scans if s.scan_type == 'arrival']
            departures = [s for s in day_scans if s.scan_type == 'departure']
            
            # Check for issues
            notes = []
            
            # Special note if vacation day but has scans (data conflict)
            if is_vacation:
                if vacation_type == 'sick_leave':
                    notes.append(f"⚠ {_('Scans on sick leave day')}")
                elif vacation_type == 'doctor':
                    notes.append(f"⚠ {_('Scans on doctor day')}")
                else:
                    notes.append(f"⚠ {_('Scans on vacation day')}")
                days_with_issues.append(current_date)
            
            # Add holiday note
            if is_holiday:
                notes.append(f"🎉 {holiday_name}")
            
            if not arrivals:
                notes.append(f"⚠ {_('Missing arrival')}")
                days_with_issues.append(current_date)
            if not departures:
                notes.append(f"⚠ {_('Missing departure')}")
                days_with_issues.append(current_date)
            
            # Calculate hours worked and lunch breaks
            hours_worked = 0
            night_hours = 0
            lunch_break_minutes = 0
            
            if arrivals and departures:
                first_arrival = arrivals[0].timestamp
                last_departure = departures[-1].timestamp
                work_duration = last_departure - first_arrival
                hours_worked = work_duration.total_seconds() / 3600
                night_hours = calculate_night_hours(first_arrival, last_departure)
                
                # Calculate lunch break
                if company.auto_lunch_breaks:
                    lunch_break_minutes = 30
                else:
                    # Calculate actual lunch break from scans
                    lunch_starts = [s for s in day_scans if s.scan_type == 'lunch_break_start']
                    lunch_ends = [s for s in day_scans if s.scan_type == 'lunch_break_end']
                    if lunch_starts and lunch_ends:
                        for i in range(min(len(lunch_starts), len(lunch_ends))):
                            break_duration = lunch_ends[i].timestamp - lunch_starts[i].timestamp
                            lunch_break_minutes += break_duration.total_seconds() / 60
                
                total_work_hours += hours_worked
                total_night_hours += night_hours
                
                # Track holiday hours
                if is_holiday:
                    total_holiday_hours += hours_worked
            
            # Get QR code info with location
            if arrivals:
                qr_info = f"{arrivals[0].qr_code.name}<br/><font size=8 color='#6b7280'>{arrivals[0].qr_code.location}</font>"
            elif departures:
                qr_info = f"{departures[0].qr_code.name}<br/><font size=8 color='#6b7280'>{departures[0].qr_code.location}</font>"
            else:
                qr_info = '-'
            
            # Format times
            arrival_time = arrivals[0].timestamp.strftime('%H:%M') if arrivals else '-'
            departure_time = departures[-1].timestamp.strftime('%H:%M') if departures else '-'
            hours_str = f"{int(hours_worked)}:{int((hours_worked % 1) * 60):02d}" if hours_worked > 0 else '0:00'
            lunch_break_str = f"{int(lunch_break_minutes)}" if lunch_break_minutes > 0 else '-'
            
            # Use conflict style for notes if vacation conflict exists
            notes_style = cell_style
            if is_vacation:
                notes_style = ParagraphStyle(
                    'ConflictNotesStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#f59e0b')
                )
            
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph(arrival_time, cell_style_centered),
                Paragraph(departure_time, cell_style_centered),
                Paragraph(hours_str, cell_style_centered),
                Paragraph(lunch_break_str, cell_style_centered),
                Paragraph(qr_info, cell_style),
                Paragraph(' '.join(notes) if notes else '✓', notes_style)
            ])
        
        current_date += timedelta(days=1)
    
    # Create table - optimized column widths (A4 landscape is 29.7cm, minus 2cm margins = 27.7cm)
    table = Table(table_data, colWidths=[2.5*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.6*cm, 1.5*cm, 7.5*cm, 7.8*cm])
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('LEFTPADDING', (0, 0), (-1, 0), 6),
        ('RIGHTPADDING', (0, 0), (-1, 0), 6),
        
        # Data rows styling
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 1), (-1, -1), 6),
        ('RIGHTPADDING', (0, 1), (-1, -1), 6),
        
        # Alignment for specific columns
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Date
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Day
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Arrival
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Departure
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Hours
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),    # Location
        ('ALIGN', (6, 1), (6, -1), 'LEFT'),    # Notes
        
        # Grid and backgrounds
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e40af')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        
        # Border styling
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6b7280')),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    summary_elements = []
    summary_elements.append(Paragraph(str(_('Summary Statistics')), heading_style))
    
    avg_hours = total_work_hours / total_days if total_days > 0 else 0
    
    # Create styled summary label and value styles
    summary_label_style = ParagraphStyle(
        'SummaryLabel',
        parent=styles['Normal'],
        fontName=font_name_bold,
        fontSize=10,
        leading=13
    )
    
    summary_value_style = ParagraphStyle(
        'SummaryValue',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10,
        leading=13
    )
    
    # Calculate overtime (based on monthly work_hours)
    # Calculate expected hours based on the date range
    months_in_range = set()
    temp_date = date_from.date()
    while temp_date <= date_to.date():
        months_in_range.add((temp_date.year, temp_date.month))
        temp_date += timedelta(days=1)
    
    # Calculate expected hours proportionally
    expected_hours = 0
    for year, month in months_in_range:
        days_in_month = monthrange(year, month)[1]
        # Days in this month that are in our range
        month_start = max(date_from.date(), datetime(year, month, 1).date())
        month_end = min(date_to.date(), datetime(year, month, days_in_month).date())
        days_in_range = (month_end - month_start).days + 1
        
        # Calculate proportional expected hours
        month_expected = (user.working_hours / days_in_month) * days_in_range
        expected_hours += month_expected
    
    overtime_hours = max(0, total_work_hours - expected_hours)
    
    # Create styled summary data
    summary_data = [
        [
            Paragraph(str(_('Total Working Days')), summary_label_style),
            Paragraph(str(total_days), summary_value_style)
        ],
        [
            Paragraph(str(_('Expected Hours')), summary_label_style),
            Paragraph(f"{int(expected_hours)}:{int((expected_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Total Hours Worked')), summary_label_style),
            Paragraph(f"{int(total_work_hours)}:{int((total_work_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Overtime Hours')), summary_label_style),
            Paragraph(f"{int(overtime_hours)}:{int((overtime_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Night Hours (22:00-06:00)')), summary_label_style),
            Paragraph(f"{int(total_night_hours)}:{int((total_night_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Holiday Hours')), summary_label_style),
            Paragraph(f"{int(total_holiday_hours)}:{int((total_holiday_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Average Hours per Day')), summary_label_style),
            Paragraph(f"{int(avg_hours)}:{int((avg_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Vacation Days')), summary_label_style),
            Paragraph(str(total_vacation_days), summary_value_style)
        ],
        [
            Paragraph(str(_('Days with Issues')), summary_label_style),
            Paragraph(str(len(set(days_with_issues))), summary_value_style)
        ],
    ]
    
    summary_table = Table(summary_data, colWidths=[10*cm, 7*cm])
    summary_table.setStyle(TableStyle([
        # Background colors
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        
        # Font styling
        ('FONTNAME', (0, 0), (0, -1), font_name_bold),
        ('FONTNAME', (1, 0), (1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        
        # Alignment
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6b7280')),
        ('LINEAFTER', (0, 0), (0, -1), 1, colors.HexColor('#a5b4fc')),
    ]))
    
    summary_elements.append(summary_table)
    
    # Add summary as a single block that won't be split across pages
    elements.append(KeepTogether(summary_elements))
    
    # Build PDF
    doc.build(elements)
    
    # Return PDF response - serve from file
    with open(filepath, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    return response


def generate_attendance_excel(request, user_id):
    """Generate Excel attendance report for a user"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from collections import defaultdict
    from calendar import monthrange
    import holidays
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    user = crud.get_user_by_id(user_id)
    if not user or user.company != company:
        messages.error(request, _('User not found'))
        return redirect('company_dashboard')
    
    # Parse date range - accept both date_range and date_from/date_to parameters
    date_range = request.GET.get('date_range', '')
    date_from_param = request.GET.get('date_from', '')
    date_to_param = request.GET.get('date_to', '')
    
    try:
        if date_from_param and date_to_param:
            # Use date_from and date_to parameters (format: YYYY-MM-DD)
            date_from = datetime.strptime(date_from_param, '%Y-%m-%d')
            date_to = datetime.strptime(date_to_param, '%Y-%m-%d')
        elif date_range and ' - ' in date_range:
            # Use date_range parameter (format: DD.MM.YYYY - DD.MM.YYYY)
            date_from_str, date_to_str = date_range.split(' - ')
            date_from = datetime.strptime(date_from_str.strip(), '%d.%m.%Y')
            date_to = datetime.strptime(date_to_str.strip(), '%d.%m.%Y')
        else:
            messages.error(request, _('Invalid date range'))
            return redirect('view_user_details', user_id=user_id)
    except:
        messages.error(request, _('Invalid date format'))
        return redirect('view_user_details', user_id=user_id)
    
    # Get scans in date range
    scans = ScanEvent.objects.filter(
        scanned_by=user,
        timestamp__date__gte=date_from.date(),
        timestamp__date__lte=date_to.date()
    ).select_related('qr_code').order_by('timestamp')
    
    vacations = Vacation.objects.filter(
        user=user,
        is_active=True,
        date_from__lte=date_to.date(),
        date_to__gte=date_from.date()
    ).order_by('date_from')
    
    # Create directory structure for Excel storage
    now = datetime.now()
    excel_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(excel_dir, exist_ok=True)
    
    # Generate filename
    filename = f"attendance_{user.name.replace(' ', '_')}_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(excel_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Attendance Report'))
    
    # Define styles
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    title_font = Font(color='2563eb', bold=True, size=16)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db')
    )
    
    # Title and header information
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{_('Attendance Report')} - {user.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{_('Period')}: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}"
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"{_('Company')}: {company.name}"
    ws['A3'].alignment = center_align
    
    # Table headers (row 5)
    headers = [
        str(_('Date')),
        str(_('Day')),
        str(_('Arrival')),
        str(_('Departure')),
        str(_('Hours')),
        str(_('Break (min)')),
        str(_('Scanned QR')),
        str(_('Notes'))
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Group scans by day
    daily_data = defaultdict(list)
    for scan in scans:
        day = scan.timestamp.date()
        daily_data[day].append(scan)
    
    # Create dictionary of vacation days with type
    vacation_days = {}
    for vacation in vacations:
        current = vacation.date_from
        while current <= vacation.date_to:
            vacation_days[current] = vacation.type if vacation.type else 'vacation'
            current += timedelta(days=1)
    
    # Helper function to calculate night hours (22:00-06:00)
    def calculate_night_hours(start_time, end_time):
        """Calculate hours worked between 22:00 and 06:00"""
        night_hours = 0
        current = start_time
        
        while current < end_time:
            # Check if current hour is night time (22:00-23:59 or 00:00-05:59)
            hour = current.hour
            if hour >= 22 or hour < 6:
                # Calculate minutes in this hour that count as night work
                next_hour = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
                segment_end = min(next_hour, end_time)
                segment_duration = (segment_end - current).total_seconds() / 3600
                night_hours += segment_duration
            
            # Move to next hour
            current = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            if current >= end_time:
                break
        
        return night_hours
    
    # Calculate statistics
    total_days = len(daily_data)
    total_work_hours = 0
    total_night_hours = 0
    days_with_issues = []
    total_vacation_days = 0
    total_holiday_hours = 0
    
    # Get holidays based on language
    lang_code = request.LANGUAGE_CODE if hasattr(request, 'LANGUAGE_CODE') else 'sk'
    if lang_code == 'sk':
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    elif lang_code == 'en':
        country_holidays = holidays.UnitedStates(years=[date_from.year, date_to.year])
    elif lang_code == 'de':
        country_holidays = holidays.Germany(years=[date_from.year, date_to.year])
    elif lang_code == 'es':
        country_holidays = holidays.Spain(years=[date_from.year, date_to.year])
    else:
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    
    # Populate daily attendance data
    row = 6
    current_date = date_from.date()
    alt_fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')
    vacation_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
    sick_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
    doctor_fill = PatternFill(start_color='ede9fe', end_color='ede9fe', fill_type='solid')
    warning_font = Font(color='f59e0b')
    
    while current_date <= date_to.date():
        day_scans = daily_data.get(current_date, [])
        day_name = date_format(current_date, format='l')
        
        # Check if this day is a holiday
        is_holiday = current_date in country_holidays
        holiday_name = country_holidays.get(current_date, '') if is_holiday else ''
        
        # Check if this day is a vacation day
        vacation_type = vacation_days.get(current_date)
        is_vacation = vacation_type is not None
        if is_vacation:
            total_vacation_days += 1
        
        # Date
        ws.cell(row=row, column=1, value=current_date.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        # Day
        ws.cell(row=row, column=2, value=day_name)
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        
        if is_vacation and not day_scans:
            # Vacation day with no scans
            ws.cell(row=row, column=3, value='-')
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value='-')
            
            if vacation_type == 'sick_leave':
                ws.cell(row=row, column=8, value=f"🏥 {_('Sick Leave')}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = sick_fill
            elif vacation_type == 'doctor':
                ws.cell(row=row, column=8, value=f"👨‍⚕️ {_('Doctor')}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = doctor_fill
            else:
                ws.cell(row=row, column=8, value=f"🏖 {_('Vacation')}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = vacation_fill
        
        elif not day_scans and not is_vacation:
            # No scans for this day
            ws.cell(row=row, column=3, value='-')
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='0:00')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value='-')
            ws.cell(row=row, column=8, value=str(_('No scans')))
        
        else:
            # Day with scans
            arrivals = [s for s in day_scans if s.scan_type == 'arrival']
            departures = [s for s in day_scans if s.scan_type == 'departure']
            
            # Check for issues
            notes = []
            
            if is_vacation:
                if vacation_type == 'sick_leave':
                    notes.append(f"⚠ {_('Scans on sick leave day')}")
                elif vacation_type == 'doctor':
                    notes.append(f"⚠ {_('Scans on doctor day')}")
                else:
                    notes.append(f"⚠ {_('Scans on vacation day')}")
                days_with_issues.append(current_date)
            
            # Add holiday note
            if is_holiday:
                notes.append(f"🎉 {holiday_name}")
            
            if not arrivals:
                notes.append(f"⚠ {_('Missing arrival')}")
                days_with_issues.append(current_date)
            if not departures:
                notes.append(f"⚠ {_('Missing departure')}")
                days_with_issues.append(current_date)
            
            # Calculate hours worked and lunch breaks
            hours_worked = 0
            night_hours = 0
            lunch_break_minutes = 0
            
            if arrivals and departures:
                first_arrival = arrivals[0].timestamp
                last_departure = departures[-1].timestamp
                work_duration = last_departure - first_arrival
                hours_worked = work_duration.total_seconds() / 3600
                night_hours = calculate_night_hours(first_arrival, last_departure)
                
                # Calculate lunch break
                if company.auto_lunch_breaks:
                    lunch_break_minutes = 30
                else:
                    # Calculate actual lunch break from scans
                    lunch_starts = [s for s in day_scans if s.scan_type == 'lunch_break_start']
                    lunch_ends = [s for s in day_scans if s.scan_type == 'lunch_break_end']
                    if lunch_starts and lunch_ends:
                        for i in range(min(len(lunch_starts), len(lunch_ends))):
                            break_duration = lunch_ends[i].timestamp - lunch_starts[i].timestamp
                            lunch_break_minutes += break_duration.total_seconds() / 60
                
                total_work_hours += hours_worked
                total_night_hours += night_hours
                
                # Track holiday hours
                if is_holiday:
                    total_holiday_hours += hours_worked
            
            # Format times
            arrival_time = arrivals[0].timestamp.strftime('%H:%M') if arrivals else '-'
            departure_time = departures[-1].timestamp.strftime('%H:%M') if departures else '-'
            hours_str = f"{int(hours_worked)}:{int((hours_worked % 1) * 60):02d}" if hours_worked > 0 else '0:00'
            lunch_break_str = str(int(lunch_break_minutes)) if lunch_break_minutes > 0 else '-'
            
            # Get QR code info
            if arrivals:
                qr_info = f"{arrivals[0].qr_code.name} - {arrivals[0].qr_code.location}"
            elif departures:
                qr_info = f"{departures[0].qr_code.name} - {departures[0].qr_code.location}"
            else:
                qr_info = '-'
            
            ws.cell(row=row, column=3, value=arrival_time)
            ws.cell(row=row, column=4, value=departure_time)
            ws.cell(row=row, column=5, value=hours_str)
            ws.cell(row=row, column=6, value=lunch_break_str)
            ws.cell(row=row, column=7, value=qr_info)
            ws.cell(row=row, column=8, value=' '.join(notes) if notes else '✓')
            
            if notes:
                ws.cell(row=row, column=8).font = warning_font
        
        # Apply alignment and borders
        for col in range(3, 9):
            if col in [3, 4, 5, 6]:  # Center align time and break columns
                ws.cell(row=row, column=col).alignment = center_align
            else:
                ws.cell(row=row, column=col).alignment = left_align
            ws.cell(row=row, column=col).border = thin_border
        
        # Alternate row colors
        if row % 2 == 0 and not is_vacation:
            for col in range(1, 9):
                if ws.cell(row=row, column=col).fill.start_color.rgb != 'd1fae5' and \
                   ws.cell(row=row, column=col).fill.start_color.rgb != 'fef3c7':
                    ws.cell(row=row, column=col).fill = alt_fill
        
        current_date += timedelta(days=1)
        row += 1
    
    # Add summary section
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value=str(_('Summary Statistics')))
    ws.cell(row=row, column=1).font = Font(color='1e40af', bold=True, size=14)
    ws.cell(row=row, column=1).alignment = left_align
    
    row += 1
    summary_fill = PatternFill(start_color='e0e7ff', end_color='e0e7ff', fill_type='solid')
    
    avg_hours = total_work_hours / total_days if total_days > 0 else 0
    
    # Calculate overtime (based on monthly work_hours)
    months_in_range = set()
    temp_date = date_from.date()
    while temp_date <= date_to.date():
        months_in_range.add((temp_date.year, temp_date.month))
        temp_date += timedelta(days=1)
    
    # Calculate expected hours proportionally
    expected_hours = 0
    for year, month in months_in_range:
        days_in_month = monthrange(year, month)[1]
        month_start = max(date_from.date(), datetime(year, month, 1).date())
        month_end = min(date_to.date(), datetime(year, month, days_in_month).date())
        days_in_range = (month_end - month_start).days + 1
        month_expected = (user.working_hours / days_in_month) * days_in_range
        expected_hours += month_expected
    
    overtime_hours = max(0, total_work_hours - expected_hours)
    
    summary_data = [
        (str(_('Total Working Days')), str(total_days)),
        (str(_('Expected Hours')), f"{int(expected_hours)}:{int((expected_hours % 1) * 60):02d}"),
        (str(_('Total Hours Worked')), f"{int(total_work_hours)}:{int((total_work_hours % 1) * 60):02d}"),
        (str(_('Overtime Hours')), f"{int(overtime_hours)}:{int((overtime_hours % 1) * 60):02d}"),
        (str(_('Night Hours (22:00-06:00)')), f"{int(total_night_hours)}:{int((total_night_hours % 1) * 60):02d}"),
        (str(_('Holiday Hours')), f"{int(total_holiday_hours)}:{int((total_holiday_hours % 1) * 60):02d}"),
        (str(_('Average Hours per Day')), f"{int(avg_hours)}:{int((avg_hours % 1) * 60):02d}"),
        (str(_('Vacation Days')), str(total_vacation_days)),
        (str(_('Days with Issues')), str(len(set(days_with_issues)))),
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=1).fill = summary_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = left_align
        
        ws.merge_cells(f'B{row}:H{row}')
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = center_align
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 25
    
    # Save workbook
    wb.save(filepath)
    
    # Return Excel response
    with open(filepath, 'rb') as excel_file:
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def generate_qr_code_pdf(request, qr_id):
    """Generate PDF with QR code for printing on A4"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        user = crud.get_user_by_id(request.session['user_id'])
        if not user or not user.is_manager or not user.can_edit_qr_codes:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    qr_code = crud.get_qr_code_by_id(qr_id)
    if not qr_code or qr_code.company != company:
        messages.error(request, _('QR Code not found'))
        return redirect('company_dashboard')
    
    # Create directory structure for PDF storage
    now = datetime.now()
    pdf_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Generate filename
    filename = f"qr_code_{qr_code.name.replace(' ', '_')}_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(pdf_dir, filename)
    
    # Register DejaVu fonts for Unicode support
    try:
        font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
        dejavu_path = os.path.join(font_dir, 'DejaVuSans.ttf')
        dejavu_bold_path = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
        
        if os.path.exists(dejavu_path) and os.path.exists(dejavu_bold_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
            font_name_bold = 'DejaVuSans-Bold'
        else:
            raise Exception("DejaVu fonts not found")
    except:
        font_name_bold = 'Helvetica-Bold'
    
    # Create PDF
    doc = SimpleDocTemplate(filepath, pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=3*cm, bottomMargin=3*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style - centered
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=24,
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    
    # Add title
    title = Paragraph(qr_code.name, title_style)
    elements.append(title)
    
    # Add spacer
    elements.append(Spacer(1, 2*cm))
    
    # Add QR code image - centered and larger
    qr_image_path = os.path.join(settings.MEDIA_ROOT, qr_code.qr_code.name)
    if os.path.exists(qr_image_path):
        # Create image with specific size (12cm x 12cm)
        img = Image(qr_image_path, width=12*cm, height=12*cm)
        img.hAlign = 'CENTER'
        elements.append(img)
    
    # Build PDF
    doc.build(elements)
    
    # Return PDF response - open in new tab
    with open(filepath, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    return response


# ============= ANALYTICS VIEWS =============

def company_analytics(request):
    """Analytics dashboard with statistics and charts"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Please login as a company or manager'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        user = crud.get_user_by_id(request.session['user_id'])
        if not user or not user.is_manager:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    from django.db.models import Count, Q, Max
    from django.utils import timezone
    from datetime import timedelta, datetime
    from calendar import monthrange
    
    today = timezone.now().date()
    
    # Get date range from request parameters
    date_from_param = request.GET.get('date_from')
    date_to_param = request.GET.get('date_to')
    
    # Parse custom date range or use defaults
    if date_from_param and date_to_param:
        try:
            date_from = datetime.strptime(date_from_param, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to_param, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
            date_to = today
    else:
        # Default: current month (from 1st to today)
        date_from = today.replace(day=1)
        date_to = today
    
    week_ago = today - timedelta(days=7)
    
    # Current month (from 1st to today) - for secondary comparison
    current_month_start = today.replace(day=1)
    
    # Previous month (full calendar month) - for secondary comparison
    if today.month == 1:
        prev_month_start = today.replace(year=today.year - 1, month=12, day=1)
        prev_month_end = today.replace(year=today.year - 1, month=12, day=31)
    else:
        prev_month_start = today.replace(month=today.month - 1, day=1)
        days_in_prev_month = monthrange(prev_month_start.year, prev_month_start.month)[1]
        prev_month_end = prev_month_start.replace(day=days_in_prev_month)
    
    # Get all users and QR codes
    users = crud.get_company_users(company)
    qr_codes = crud.get_company_qr_codes(company)
    
    # Today's statistics (always based on actual today)
    today_scans = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date=today
    )
    today_arrivals = today_scans.filter(scan_type='arrival').count()
    today_departures = today_scans.filter(scan_type='departure').count()
    
    # Statistics for selected date range
    range_scans = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date__gte=date_from,
        timestamp__date__lte=date_to
    )
    range_arrivals = range_scans.filter(scan_type='arrival').count()
    range_departures = range_scans.filter(scan_type='departure').count()
    range_total_scans = range_scans.count()
    
    # Weekly statistics
    week_scans = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date__gte=week_ago
    ).count()
    
    # Current month statistics
    current_month_scans = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date__gte=current_month_start,
        timestamp__date__lte=today
    ).count()
    
    # Previous month statistics
    prev_month_scans = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date__gte=prev_month_start,
        timestamp__date__lte=prev_month_end
    ).count()
    
    # Currently in office (last scan was arrival)
    currently_in_office = []
    for user in users:
        last_scan = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__company=company
        ).order_by('-timestamp').first()
        
        if last_scan and last_scan.scan_type == 'arrival':
            currently_in_office.append({
                'user': user,
                'location': last_scan.qr_code.name,
                'time': last_scan.timestamp
            })
    
    # Top 5 most used QR codes (for selected date range)
    top_qr_codes = ScanEvent.objects.filter(
        qr_code__company=company,
        timestamp__date__gte=date_from,
        timestamp__date__lte=date_to
    ).values('qr_code__name', 'qr_code__location').annotate(
        scan_count=Count('id')
    ).order_by('-scan_count')[:5]
    
    # Calculate working hours for selected date range
    selected_range_work_hours = []
    for user in users:
        scans = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__company=company,
            timestamp__date__gte=date_from,
            timestamp__date__lte=date_to
        ).order_by('timestamp')
        
        total_hours = 0
        arrival_time = None
        
        for scan in scans:
            if scan.scan_type == 'arrival':
                arrival_time = scan.timestamp
            elif scan.scan_type == 'departure' and arrival_time:
                work_duration = (scan.timestamp - arrival_time).total_seconds() / 3600
                total_hours += work_duration
                arrival_time = None
        
        if total_hours > 0:  # Only include users with hours
            selected_range_work_hours.append({
                'user': user,
                'hours': round(total_hours, 1),
                'days': scans.values('timestamp__date').distinct().count()
            })
    
    # Calculate working hours for current month (calendar)
    current_month_work_hours = []
    for user in users:
        scans = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__company=company,
            timestamp__date__gte=current_month_start,
            timestamp__date__lte=today
        ).order_by('timestamp')
        
        total_hours = 0
        arrival_time = None
        
        for scan in scans:
            if scan.scan_type == 'arrival':
                arrival_time = scan.timestamp
            elif scan.scan_type == 'departure' and arrival_time:
                work_duration = (scan.timestamp - arrival_time).total_seconds() / 3600
                total_hours += work_duration
                arrival_time = None
        
        if total_hours > 0:  # Only include users with hours
            current_month_work_hours.append({
                'user': user,
                'hours': round(total_hours, 1),
                'days': scans.values('timestamp__date').distinct().count()
            })
    
    # Calculate working hours for previous month
    prev_month_work_hours = []
    for user in users:
        scans = ScanEvent.objects.filter(
            scanned_by=user,
            qr_code__company=company,
            timestamp__date__gte=prev_month_start,
            timestamp__date__lte=prev_month_end
        ).order_by('timestamp')
        
        total_hours = 0
        arrival_time = None
        
        for scan in scans:
            if scan.scan_type == 'arrival':
                arrival_time = scan.timestamp
            elif scan.scan_type == 'departure' and arrival_time:
                work_duration = (scan.timestamp - arrival_time).total_seconds() / 3600
                total_hours += work_duration
                arrival_time = None
        
        if total_hours > 0:  # Only include users with hours
            prev_month_work_hours.append({
                'user': user,
                'hours': round(total_hours, 1),
                'days': scans.values('timestamp__date').distinct().count()
            })
    
    # Sort by hours
    selected_range_work_hours.sort(key=lambda x: x['hours'], reverse=True)
    current_month_work_hours.sort(key=lambda x: x['hours'], reverse=True)
    prev_month_work_hours.sort(key=lambda x: x['hours'], reverse=True)
    
    # Active vacations in selected range
    active_vacations = Vacation.objects.filter(
        user__company=company,
        is_active=True,
        date_from__lte=date_to,
        date_to__gte=date_from
    ).count()
    
    # Calculate number of days in selected range
    days_in_range = (date_to - date_from).days + 1
    
    context = {
        'company': company,
        'today_arrivals': today_arrivals,
        'today_departures': today_departures,
        'week_scans': week_scans,
        'current_month_scans': current_month_scans,
        'prev_month_scans': prev_month_scans,
        'current_month_name': current_month_start.strftime('%B %Y'),
        'prev_month_name': prev_month_start.strftime('%B %Y'),
        'currently_in_office': currently_in_office,
        'currently_in_office_count': len(currently_in_office),
        'top_qr_codes': top_qr_codes,
        'selected_range_work_hours': selected_range_work_hours,
        'current_month_work_hours': current_month_work_hours,
        'prev_month_work_hours': prev_month_work_hours,
        'total_users': users.count(),
        'total_qr_codes': qr_codes.count(),
        'active_vacations': active_vacations,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'date_range_display': f"{date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}",
        'range_arrivals': range_arrivals,
        'range_departures': range_departures,
        'range_total_scans': range_total_scans,
        'days_in_range': days_in_range,
    }
    
    return render(request, 'company_analytics.html', context)


def analytics_chart_data(request):
    """API endpoint for chart data (JSON)"""
    if 'company_id' not in request.session or request.session.get('user_type') != 'company':
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    company = crud.get_company_by_id(request.session['company_id'])
    if not company:
        return JsonResponse({'error': 'Company not found'}, status=404)
    
    from django.db.models import Count
    from django.utils import timezone
    from datetime import timedelta
    
    chart_type = request.GET.get('type', 'daily')
    days = int(request.GET.get('days', 7))
    
    today = timezone.now().date()
    start_date = today - timedelta(days=days-1)
    
    if chart_type == 'daily':
        # Daily scan counts
        data = []
        labels = []
        arrivals_data = []
        departures_data = []
        
        for i in range(days):
            date = start_date + timedelta(days=i)
            labels.append(date.strftime('%d.%m'))
            
            day_scans = ScanEvent.objects.filter(
                qr_code__company=company,
                timestamp__date=date
            )
            
            arrivals = day_scans.filter(scan_type='arrival').count()
            departures = day_scans.filter(scan_type='departure').count()
            
            arrivals_data.append(arrivals)
            departures_data.append(departures)
        
        return JsonResponse({
            'labels': labels,
            'datasets': [
                {
                    'label': str(_('Arrivals')),
                    'data': arrivals_data,
                    'borderColor': 'rgb(16, 185, 129)',
                    'backgroundColor': 'rgba(16, 185, 129, 0.1)',
                    'tension': 0.4
                },
                {
                    'label': str(_('Departures')),
                    'data': departures_data,
                    'borderColor': 'rgb(239, 68, 68)',
                    'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                    'tension': 0.4
                }
            ]
        })
    
    elif chart_type == 'qr_usage':
        # QR code usage pie chart
        qr_data = ScanEvent.objects.filter(
            qr_code__company=company,
            timestamp__date__gte=start_date
        ).values('qr_code__name').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        labels = [item['qr_code__name'] for item in qr_data]
        data = [item['count'] for item in qr_data]
        
        colors = [
            'rgb(37, 99, 235)',
            'rgb(16, 185, 129)',
            'rgb(245, 158, 11)',
            'rgb(239, 68, 68)',
            'rgb(139, 92, 246)'
        ]
        
        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'data': data,
                'backgroundColor': colors[:len(data)],
                'borderWidth': 2,
                'borderColor': '#fff'
            }]
        })
    
    elif chart_type == 'hourly':
        # Hourly distribution (today)
        hours = list(range(24))
        labels = [f'{h:02d}:00' for h in hours]
        data = []
        
        for hour in hours:
            count = ScanEvent.objects.filter(
                qr_code__company=company,
                timestamp__date=today,
                timestamp__hour=hour
            ).count()
            data.append(count)
        
        return JsonResponse({
            'labels': labels,
            'datasets': [{
                'label': str(_('Scans')),
                'data': data,
                'backgroundColor': 'rgba(37, 99, 235, 0.5)',
                'borderColor': 'rgb(37, 99, 235)',
                'borderWidth': 2
            }]
        })
    
    return JsonResponse({'error': 'Invalid chart type'}, status=400)


# ============= AUDIT LOGGING VIEWS =============

def audit_logs(request):
    """View audit logs - company and managers see all, regular users see only their own"""
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_user = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_user):
        messages.error(request, _('Please login to access this page'))
        return redirect('landing_page')
    
    from viewer.models import AuditLog
    
    # Get company and current user
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
        current_user = None
        is_manager = False
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user:
            messages.error(request, _('User not found'))
            return redirect('user_login')
        company = current_user.company
        is_manager = current_user.is_manager
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('landing_page')
    
    # Get all logs - filter based on user type
    if is_company or is_manager:
        # Company and managers see all logs for their company
        # Get all user emails from company
        users = company.users.filter(is_active=True).values_list('email', flat=True)
        logs = AuditLog.objects.filter(
            Q(actor_email=company.email) | Q(actor_email__in=users)
        )
    else:
        # Regular users see only their own logs
        logs = AuditLog.objects.filter(actor_email=current_user.email)
    
    # Filtering
    actor_filter = request.GET.get('actor', '')
    action_filter = request.GET.get('action', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if actor_filter:
        logs = logs.filter(Q(actor_name__icontains=actor_filter) | Q(actor_email__icontains=actor_filter))
    
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    if date_from and date_to:
        from datetime import datetime
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            logs = logs.filter(timestamp__date__gte=date_from_obj.date(), timestamp__date__lte=date_to_obj.date())
        except:
            messages.error(request, _('Invalid date range'))
    
    # Sorting
    sort = request.GET.get('sort', '-timestamp')
    valid_sort_fields = ['timestamp', '-timestamp', 'actor_name', '-actor_name', 'action', '-action']
    
    if sort in valid_sort_fields:
        logs = logs.order_by(sort)
    else:
        logs = logs.order_by('-timestamp')
    
    # Pagination
    items_per_page = request.GET.get('items_per_page', '25')
    try:
        items_per_page = int(items_per_page)
        if items_per_page not in [10, 25, 50, 100]:
            items_per_page = 25
    except (ValueError, TypeError):
        items_per_page = 25
    
    page_number = request.GET.get('page', 1)
    paginator = Paginator(logs, items_per_page)
    page_obj = paginator.get_page(page_number)
    
    # Get unique values for filters
    if is_company or is_manager:
        all_logs = AuditLog.objects.filter(
            Q(actor_email=company.email) | Q(actor_email__in=users)
        )
    else:
        all_logs = AuditLog.objects.filter(actor_email=current_user.email)
    
    unique_actors = all_logs.values_list('actor_name', flat=True).distinct()
    unique_actions = all_logs.values_list('action', flat=True).distinct()
    
    context = {
        'company': company,
        'current_user': current_user,
        'is_company': is_company,
        'is_manager': is_manager,
        'page_obj': page_obj,
        'logs_count': logs.count(),
        'unique_actors': list(unique_actors),
        'unique_actions': list(unique_actions),
        'current_filters': {
            'actor': actor_filter,
            'action': action_filter,
            'date_from': date_from,
            'date_to': date_to,
            'sort': sort,
            'items_per_page': str(items_per_page),
        }
    }
    return render(request, 'audit_logs.html', context)


# ============= MAGAZINE VIEWS =============

def magazine_dashboard(request):
    """Magazine dashboard - list all magazines for the company"""
    if 'company_id' not in request.session or request.session.get('user_type') != 'company':
        messages.error(request, _('Please login as a company to access this page'))
        return redirect('company_login')
    
    from viewer.models import Magazine
    company_id = request.session['company_id']
    magazines = Magazine.objects.filter(company_id=company_id).order_by('-modified_at')
    
    context = {
        'magazines': magazines,
        'company_id': company_id,
    }
    return render(request, 'magazine_dashboard.html', context)


def magazine_editor(request, magazine_id=None):
    """Magazine editor - create or edit a magazine"""
    if 'company_id' not in request.session or request.session.get('user_type') != 'company':
        messages.error(request, _('Please login as a company to access this page'))
        return redirect('company_login')
    
    from viewer.models import Magazine, MagazineArticle, User
    company_id = request.session['company_id']
    
    # Get or create magazine
    if magazine_id:
        magazine = Magazine.objects.filter(id=magazine_id, company_id=company_id).first()
        if not magazine:
            messages.error(request, _('Magazine not found'))
            return redirect('magazine_dashboard')
    else:
        # Create new magazine
        user = User.objects.filter(company_id=company_id).first()
        magazine = Magazine.objects.create(
            company_id=company_id,
            created_by=user,
            title="New Magazine",
            issue_number="1",
            publish_date=datetime.date.today()
        )
        
        MagazineArticle.objects.create(
            magazine=magazine,
            author=user,
            title="Article 1",
            category="News"
        )
        
        return redirect('magazine_editor', magazine_id=magazine.id)
    
    # Get articles
    articles = magazine.articles.all()
    users = User.objects.filter(company_id=company_id, is_active=True)
    
    context = {
        'magazine': magazine,
        'articles': articles,
        'users': users,
        'categories': magazine.get_categories_list(),
    }
    return render(request, 'magazine_editor.html', context)


def magazine_preview(request, magazine_id):
    """Magazine preview - show print-ready preview"""
    if 'company_id' not in request.session or request.session.get('user_type') != 'company':
        messages.error(request, _('Please login as a company to access this page'))
        return redirect('company_login')
    
    from viewer.models import Magazine
    company_id = request.session['company_id']
    
    magazine = Magazine.objects.filter(id=magazine_id, company_id=company_id).first()
    if not magazine:
        messages.error(request, _('Magazine not found'))
        return redirect('magazine_dashboard')
    
    articles = magazine.articles.all().order_by('order', 'page_number')
    
    context = {
        'magazine': magazine,
        'articles': articles,
    }
    return render(request, 'magazine_preview.html', context)


# ============= MAGAZINE API ENDPOINTS =============

@csrf_exempt
def api_magazine_update(request, magazine_id):
    """API: Update magazine configuration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import Magazine
    company_id = request.session['company_id']
    
    magazine = Magazine.objects.filter(id=magazine_id, company_id=company_id).first()
    if not magazine:
        return JsonResponse({'error': 'Magazine not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        
        # Update fields
        if 'title' in data:
            magazine.title = data['title']
        if 'issue_number' in data:
            magazine.issue_number = data['issue_number']
        if 'tagline' in data:
            magazine.tagline = data['tagline']
        if 'publish_date' in data:
            magazine.publish_date = data['publish_date']
        if 'template_id' in data:
            magazine.template_id = data['template_id']
        if 'primary_color' in data:
            magazine.primary_color = data['primary_color']
        if 'secondary_color' in data:
            magazine.secondary_color = data['secondary_color']
        if 'background_color' in data:
            magazine.background_color = data['background_color']
        if 'categories' in data:
            magazine.categories = data['categories']
        if 'cover_background_image' in data:
            magazine.cover_background_image = data['cover_background_image']
        if 'cover_header_position' in data:
            magazine.cover_header_position = data['cover_header_position']
        if 'primary_font' in data:
            magazine.primary_font = data['primary_font']
        if 'secondary_font' in data:
            magazine.secondary_font = data['secondary_font']
        if 'text_color' in data:
            magazine.text_color = data['text_color']
        
        magazine.save()
        
        # Log the update
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='update',
                message=f'Magazine "{magazine.title}" updated',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True, 'magazine': {
            'id': magazine.id,
            'title': magazine.title,
            'issue_number': magazine.issue_number,
        }})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_magazine_delete(request, magazine_id):
    """API: Delete a magazine"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import Magazine
    company_id = request.session['company_id']
    
    magazine = Magazine.objects.filter(id=magazine_id, company_id=company_id).first()
    if not magazine:
        return JsonResponse({'error': 'Magazine not found'}, status=404)
    
    magazine_title = magazine.title
    magazine.delete()
    
    # Log the deletion
    company = crud.get_company_by_id(company_id)
    if company:
        log_action(
            actor_type='company',
            actor_email=company.email,
            actor_name=company.name,
            action='delete',
            message=f'Magazine "{magazine_title}" deleted',
            ip_address=get_client_ip(request)
        )
    
    return JsonResponse({'success': True})


@csrf_exempt
def api_article_create(request, magazine_id):
    """API: Create a new article"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import Magazine, MagazineArticle, User
    company_id = request.session['company_id']
    
    magazine = Magazine.objects.filter(id=magazine_id, company_id=company_id).first()
    if not magazine:
        return JsonResponse({'error': 'Magazine not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        
        # Get author
        user = User.objects.filter(company_id=company_id).first()
        
        article = MagazineArticle.objects.create(
            magazine=magazine,
            author=user,
            title=data.get('title', 'New Article'),
            category=data.get('category', magazine.get_categories_list()[0])
        )
        
        # Log the creation
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='create',
                message=f'Article "{article.title}" created in magazine "{magazine.title}"',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True, 'article': {
            'id': article.id,
            'title': article.title,
            'category': article.category,
        }})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_article_update(request, article_id):
    """API: Update an article"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        
        if 'title' in data:
            article.title = data['title']
        if 'teaser' in data:
            article.teaser = data['teaser']
        if 'category' in data:
            article.category = data['category']
        if 'is_main_story' in data:
            article.is_main_story = data['is_main_story']
        if 'is_secondary_story' in data:
            article.is_secondary_story = data['is_secondary_story']
        if 'order' in data:
            article.order = data['order']
        if 'status' in data:
            article.status = data['status']
        
        article.save()
        
        # Log the update
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='update',
                message=f'Article "{article.title}" updated',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True, 'article': {
            'id': article.id,
            'title': article.title,
            'status': article.status,
        }})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_article_delete(request, article_id):
    """API: Delete an article"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    article_title = article.title
    article.delete()
    
    # Log the deletion
    company = crud.get_company_by_id(company_id)
    if company:
        log_action(
            actor_type='company',
            actor_email=company.email,
            actor_name=company.name,
            action='delete',
            message=f'Article "{article_title}" deleted',
            ip_address=get_client_ip(request)
        )
    
    return JsonResponse({'success': True})


@csrf_exempt
def api_article_upload_header_image(request, article_id):
    """API: Upload header image for article"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    try:
        if not request.FILES.get('header_image'):
            return JsonResponse({'error': 'No image provided'}, status=400)
        
        article.header_image = request.FILES['header_image']
        article.save()
        
        # Log the upload
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='update',
                message=f'Header image uploaded for article "{article.title}"',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True, 'header_image': article.header_image.url})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_article_remove_header_image(request, article_id):
    """API: Remove header image from article"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    try:
        if article.header_image:
            article.header_image.delete()
        article.header_image = None
        article.save()
        
        # Log the removal
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='update',
                message=f'Header image removed from article "{article.title}"',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_content_block_create(request, article_id):
    """API: Create a content block"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle, ContentBlock
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    try:
        # Get next order
        max_order = article.content_blocks.aggregate(Max('order'))['order__max'] or 0
        
        # Check if it's a file upload (image)
        if request.FILES.get('image'):
            block = ContentBlock.objects.create(
                article=article,
                block_type='image',
                order=max_order + 1,
                image=request.FILES['image'],
                alignment='center'
            )
        else:
            # JSON data
            data = json.loads(request.body)
            
            block = ContentBlock.objects.create(
                article=article,
                block_type=data.get('block_type', 'text'),
                order=max_order + 1,
                text_content=data.get('text_content', ''),
                image_url=data.get('image_url', ''),
                alignment=data.get('alignment', 'left')
            )
        
        # Log the creation
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='create',
                message=f'Content block ({block.block_type}) created in article "{article.title}"',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True, 'block': {
            'id': block.id,
            'block_type': block.block_type,
            'order': block.order,
        }})
    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=400)


@csrf_exempt
def api_content_block_update(request, block_id):
    """API: Update a content block"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import ContentBlock
    company_id = request.session['company_id']
    
    block = ContentBlock.objects.filter(
        id=block_id,
        article__magazine__company_id=company_id
    ).first()
    
    if not block:
        return JsonResponse({'error': 'Block not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        
        if 'text_content' in data:
            block.text_content = data['text_content']
        if 'image_url' in data:
            block.image_url = data['image_url']
        if 'image_caption' in data:
            block.image_caption = data['image_caption']
        if 'alignment' in data:
            block.alignment = data['alignment']
        if 'font_size' in data:
            block.font_size = data['font_size']
        if 'order' in data:
            block.order = data['order']
        
        block.save()
        
        # Log the update
        company = crud.get_company_by_id(company_id)
        if company:
            log_action(
                actor_type='company',
                actor_email=company.email,
                actor_name=company.name,
                action='update',
                message=f'Content block updated in article "{block.article.title}"',
                ip_address=get_client_ip(request)
            )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def api_content_block_delete(request, block_id):
    """API: Delete a content block"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import ContentBlock
    company_id = request.session['company_id']
    
    block = ContentBlock.objects.filter(
        id=block_id,
        article__magazine__company_id=company_id
    ).first()
    
    if not block:
        return JsonResponse({'error': 'Block not found'}, status=404)
    
    article_title = block.article.title
    block.delete()
    
    # Log the deletion
    company = crud.get_company_by_id(company_id)
    if company:
        log_action(
            actor_type='company',
            actor_email=company.email,
            actor_name=company.name,
            action='delete',
            message=f'Content block deleted from article "{article_title}"',
            ip_address=get_client_ip(request)
        )
    
    return JsonResponse({'success': True})


@csrf_exempt
def api_article_reorder_blocks(request, article_id):
    """API: Reorder content blocks"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle, ContentBlock
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    try:
        data = json.loads(request.body)
        blocks = data.get('blocks', [])
        
        # Update each block's order
        for block_data in blocks:
            block = ContentBlock.objects.filter(
                id=block_data['id'],
                article=article
            ).first()
            if block:
                block.order = block_data['order']
                block.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def api_article_data(request, article_id):
    """API: Get article data with content blocks"""
    if 'company_id' not in request.session:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    
    from viewer.models import MagazineArticle
    company_id = request.session['company_id']
    
    article = MagazineArticle.objects.filter(
        id=article_id,
        magazine__company_id=company_id
    ).first()
    
    if not article:
        return JsonResponse({'error': 'Article not found'}, status=404)
    
    # Get content blocks
    blocks = []
    for block in article.content_blocks.all():
        block_data = {
            'id': block.id,
            'block_type': block.block_type,
            'order': block.order,
            'alignment': block.alignment,
        }
        
        if block.block_type == 'text':
            block_data['text_content'] = block.text_content
            block_data['font_size'] = block.font_size
        elif block.block_type == 'image':
            block_data['image_url'] = block.image_url
            block_data['image_caption'] = block.image_caption
            if block.image:
                block_data['image'] = block.image.url
        
        blocks.append(block_data)
    
    return JsonResponse({
        'success': True,
        'article': {
            'id': article.id,
            'title': article.title,
            'teaser': article.teaser,
            'category': article.category,
            'is_main_story': article.is_main_story,
            'is_secondary_story': article.is_secondary_story,
            'status': article.status,
            'header_image': article.header_image.url if article.header_image else None,
            'content_blocks': blocks
        }
    })
