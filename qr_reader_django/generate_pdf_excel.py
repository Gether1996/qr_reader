from django.shortcuts import redirect
from django.contrib import messages
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.conf import settings
from qr_reader_django import crud
from viewer.models import ScanEvent, Vacation
from django.utils.formats import date_format

def generate_attendance_pdf(request, user_id):
    """Generate PDF attendance report for a user"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime, timedelta
    from collections import defaultdict
    from calendar import monthrange
    import holidays
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    user = crud.get_user_by_id(user_id, include_inactive=True)
    if not user or user.company != company:
        messages.error(request, _('User not found'))
        return redirect('company_dashboard')
    
    # Parse date range - accept both date_range and date_from/date_to parameters
    date_range = request.GET.get('date_range', '')
    date_from_param = request.GET.get('date_from', '')
    date_to_param = request.GET.get('date_to', '')
    
    try:
        if date_from_param and date_to_param:
            # Use date_from and date_to parameters (format: YYYY-MM-DD)
            date_from = datetime.strptime(date_from_param, '%Y-%m-%d')
            date_to = datetime.strptime(date_to_param, '%Y-%m-%d')
        elif date_range and ' - ' in date_range:
            # Use date_range parameter (format: DD.MM.YYYY - DD.MM.YYYY)
            date_from_str, date_to_str = date_range.split(' - ')
            date_from = datetime.strptime(date_from_str.strip(), '%d.%m.%Y')
            date_to = datetime.strptime(date_to_str.strip(), '%d.%m.%Y')
        else:
            messages.error(request, _('Invalid date range'))
            return redirect('view_user_details', user_id=user_id)
    except:
        messages.error(request, _('Invalid date format'))
        return redirect('view_user_details', user_id=user_id)
    
    # Get scans in date range
    scans = ScanEvent.objects.filter(
        scanned_by=user,
        timestamp__date__gte=date_from.date(),
        timestamp__date__lte=date_to.date()
    ).select_related('qr_code').order_by('timestamp')
    
    vacations = Vacation.objects.filter(
        user=user,
        is_active=True,
        date_from__lte=date_to.date(),
        date_to__gte=date_from.date()
    ).order_by('date_from')
    
    # Create directory structure for PDF storage
    now = datetime.now()
    pdf_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Generate filename
    filename = f"attendance_{user.name.replace(' ', '_')}_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.pdf"
    filepath = os.path.join(pdf_dir, filename)
    
    # Register DejaVu fonts for Unicode support (Slovak characters)
    try:
        # Use fonts from project directory
        font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
        dejavu_path = os.path.join(font_dir, 'DejaVuSans.ttf')
        dejavu_bold_path = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
        
        if os.path.exists(dejavu_path) and os.path.exists(dejavu_bold_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
            font_name = 'DejaVuSans'
            font_name_bold = 'DejaVuSans-Bold'
        else:
            raise Exception("DejaVu fonts not found in project")
    except Exception as e:
        # Fallback to Helvetica if DejaVu is not available
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
    
    # Create PDF
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), 
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Unicode-compatible font
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=20,
        textColor=colors.HexColor('#2563eb'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_name_bold,
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=11
    )
    
    # Title
    elements.append(Paragraph(f"{_('Attendance Report')} - {user.name}", title_style))
    elements.append(Paragraph(f"{_('Period')}: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}", normal_style))
    elements.append(Paragraph(f"{_('Company')}: {company.name}", normal_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Group scans by day
    daily_data = defaultdict(list)
    for scan in scans:
        day = scan.timestamp.date()
        daily_data[day].append(scan)
    
    # Create dictionary of vacation days with type
    vacation_days = {}
    for vacation in vacations:
        current = vacation.date_from
        while current <= vacation.date_to:
            vacation_days[current] = vacation.type if vacation.type else 'vacation'
            current += timedelta(days=1)
    
    # Helper function to calculate night hours (22:00-06:00)
    def calculate_night_hours(start_time, end_time):
        """Calculate hours worked between 22:00 and 06:00"""
        night_hours = 0
        current = start_time
        
        while current < end_time:
            # Check if current hour is night time (22:00-23:59 or 00:00-05:59)
            hour = current.hour
            if hour >= 22 or hour < 6:
                # Calculate minutes in this hour that count as night work
                next_hour = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
                segment_end = min(next_hour, end_time)
                segment_duration = (segment_end - current).total_seconds() / 3600
                night_hours += segment_duration
            
            # Move to next hour
            current = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            if current >= end_time:
                break
        
        return night_hours
    
    # Calculate statistics
    total_days = len(daily_data)
    total_work_hours = 0
    total_work_hours_with_breaks = 0
    total_night_hours = 0
    total_break_minutes = 0
    days_with_issues = []
    total_vacation_days = 0
    total_home_office_days = 0
    
    # Daily attendance table
    elements.append(Paragraph(str(_('Daily Attendance')), heading_style))
    
    # Use Paragraph objects for header cells to enable wrapping
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName=font_name_bold,
        fontSize=9,
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
        leading=11
    )
    
    table_data = [[
        Paragraph(str(_('Date')), header_style),
        Paragraph(str(_('Day')), header_style),
        Paragraph(str(_('Arrival')), header_style),
        Paragraph(str(_('Departure')), header_style),
        Paragraph(str(_('Hours')), header_style),
        Paragraph(str(_('Break')), header_style),
        Paragraph(str(_('Scanned QR')), header_style),
        Paragraph(str(_('Notes')), header_style)
    ]]
    
    # Style for data cells
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9,
        leading=11
    )
    
    cell_style_centered = ParagraphStyle(
        'CellStyleCentered',
        parent=cell_style,
        alignment=TA_CENTER
    )
    
    # Get holidays based on language
    lang_code = request.LANGUAGE_CODE if hasattr(request, 'LANGUAGE_CODE') else 'sk'
    if lang_code == 'sk':
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    elif lang_code == 'en':
        country_holidays = holidays.UnitedStates(years=[date_from.year, date_to.year])
    elif lang_code == 'de':
        country_holidays = holidays.Germany(years=[date_from.year, date_to.year])
    elif lang_code == 'es':
        country_holidays = holidays.Spain(years=[date_from.year, date_to.year])
    else:
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    
    total_holiday_hours = 0
    
    current_date = date_from.date()
    while current_date <= date_to.date():
        day_scans = daily_data.get(current_date, [])
        # Use Django's date_format with 'l' format (day of the week)
        day_name = date_format(current_date, format='l')
        
        # Check if this day is a holiday
        is_holiday = current_date in country_holidays
        holiday_name = country_holidays.get(current_date, '') if is_holiday else ''
        
        # Check if this day is a vacation day
        vacation_obj = vacation_days.get(current_date)
        is_vacation = vacation_obj is not None
        vacation_type = vacation_obj.type if vacation_obj else None
        if is_vacation:
            if vacation_type == 'home_office':
                total_home_office_days += 1
            else:
                total_vacation_days += 1
        
        if is_vacation and not day_scans:
            # Vacation day with no scans - display type
            # Format time if available
            time_info = ""
            if vacation_obj and vacation_obj.time_from and vacation_obj.time_to:
                time_info = f" ({vacation_obj.time_from.strftime('%H:%M')} - {vacation_obj.time_to.strftime('%H:%M')})"
            
            if vacation_type == 'sick_leave':
                vacation_style = ParagraphStyle(
                    'SickLeaveStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#f59e0b'),
                    fontName=font_name_bold
                )
                leave_label = f"🏥 {_('Sick Leave')}{time_info}"
            elif vacation_type == 'doctor':
                vacation_style = ParagraphStyle(
                    'DoctorStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#8b5cf6'),
                    fontName=font_name_bold
                )
                leave_label = f"👨‍⚕️ {_('Doctor')}{time_info}"
            elif vacation_type == 'home_office':
                vacation_style = ParagraphStyle(
                    'HomeOfficeStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#3b82f6'),
                    fontName=font_name_bold
                )
                leave_label = f"🏠 {_('Home Office')}{time_info}"
            else:
                vacation_style = ParagraphStyle(
                    'VacationStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#10b981'),
                    fontName=font_name_bold
                )
                leave_label = f"🏖 {_('Vacation')}{time_info}"
            
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style),
                Paragraph(leave_label, vacation_style)
            ])
        elif not day_scans and not is_vacation:
            # No scans for this day
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('0:00', cell_style_centered),
                Paragraph('-', cell_style_centered),
                Paragraph('-', cell_style),
                Paragraph(str(_('No scans')), cell_style)
            ])
        else:
            # Day with scans (may or may not be vacation)
            # Find arrivals and departures
            arrivals = [s for s in day_scans if s.scan_type == 'arrival']
            departures = [s for s in day_scans if s.scan_type == 'departure']
            
            # Check for issues
            notes = []
            
            # Special note if vacation day but has scans (data conflict)
            if is_vacation:
                if vacation_type == 'sick_leave':
                    notes.append(f"⚠ {_('Scans on sick leave day')}")
                elif vacation_type == 'doctor':
                    notes.append(f"⚠ {_('Scans on doctor day')}")
                elif vacation_type == 'home_office':
                    notes.append(f"🏠 {_('Home Office')}")
                else:
                    notes.append(f"⚠ {_('Scans on vacation day')}")
                # Only mark as issue if not home office (home office + scans is expected)
                if vacation_type != 'home_office':
                    days_with_issues.append(current_date)
            
            # Add holiday note
            if is_holiday:
                notes.append(f"🎉 {holiday_name}")
            
            if not arrivals:
                notes.append(f"⚠ {_('Missing arrival')}")
                days_with_issues.append(current_date)
            if not departures:
                notes.append(f"⚠ {_('Missing departure')}")
                days_with_issues.append(current_date)
            
            # Calculate hours worked and lunch breaks
            hours_worked = 0
            night_hours = 0
            lunch_break_minutes = 0
            
            if arrivals and departures:
                first_arrival = arrivals[0].timestamp
                last_departure = departures[-1].timestamp
                work_duration = last_departure - first_arrival
                hours_worked = work_duration.total_seconds() / 3600
                night_hours = calculate_night_hours(first_arrival, last_departure)
                
                # Calculate lunch break
                # First check if there are actual lunch break scans
                lunch_starts = [s for s in day_scans if s.scan_type == 'lunch_break_start']
                lunch_ends = [s for s in day_scans if s.scan_type == 'lunch_break_end']
                
                if lunch_starts and lunch_ends:
                    # Use actual scanned lunch breaks
                    for i in range(min(len(lunch_starts), len(lunch_ends))):
                        break_duration = lunch_ends[i].timestamp - lunch_starts[i].timestamp
                        lunch_break_minutes += break_duration.total_seconds() / 60
                elif company.auto_lunch_breaks and user.has_lunch_break:
                    # No scanned breaks, use default duration if auto breaks enabled
                    lunch_break_minutes = user.lunch_break_duration
                # else: lunch_break_minutes stays 0
                
                # Track hours with and without breaks
                hours_without_break = hours_worked - (lunch_break_minutes / 60)
                total_work_hours += hours_without_break  # Hours without break
                total_work_hours_with_breaks += hours_worked  # Hours including break
                total_night_hours += night_hours
                
                # Track holiday hours
                if is_holiday:
                    total_holiday_hours += hours_worked
            
            # Get QR code info with location
            if arrivals:
                if arrivals[0].is_home_office:
                    qr_info = f"{_('Home Office')}<br/><font size=8 color='#6b7280'>{_('Home Office')}</font>"
                elif arrivals[0].is_business_trip:
                    qr_info = f"{_('Business Trip')}<br/><font size=8 color='#6b7280'>{_('Business Trip')}</font>"
                else:
                    qr_info = f"{arrivals[0].qr_code.name}<br/><font size=8 color='#6b7280'>{arrivals[0].qr_code.location}</font>"
            elif departures:
                if departures[0].is_home_office:
                    qr_info = f"{_('Home Office')}<br/><font size=8 color='#6b7280'>{_('Home Office')}</font>"
                elif departures[0].is_business_trip:
                    qr_info = f"{_('Business Trip')}<br/><font size=8 color='#6b7280'>{_('Business Trip')}</font>"
                else:
                    qr_info = f"{departures[0].qr_code.name}<br/><font size=8 color='#6b7280'>{departures[0].qr_code.location}</font>"
            else:
                qr_info = '-'
            
            # Format times
            arrival_time = arrivals[0].timestamp.strftime('%H:%M') if arrivals else '-'
            departure_time = departures[-1].timestamp.strftime('%H:%M') if departures else '-'
            hours_str = f"{int(hours_worked)}:{int((hours_worked % 1) * 60):02d}" if hours_worked > 0 else '0:00'
            # Format break time as HH:MM
            if lunch_break_minutes > 0:
                break_hours = int(lunch_break_minutes // 60)
                break_mins = int(lunch_break_minutes % 60)
                lunch_break_str = f"{break_hours}:{break_mins:02d}"
                total_break_minutes += lunch_break_minutes
            else:
                lunch_break_str = '-'
            
            # Use conflict style for notes if vacation conflict exists
            notes_style = cell_style
            if is_vacation:
                notes_style = ParagraphStyle(
                    'ConflictNotesStyle',
                    parent=cell_style,
                    textColor=colors.HexColor('#f59e0b')
                )
            
            table_data.append([
                Paragraph(current_date.strftime('%d.%m.%Y'), cell_style_centered),
                Paragraph(day_name, cell_style),
                Paragraph(arrival_time, cell_style_centered),
                Paragraph(departure_time, cell_style_centered),
                Paragraph(hours_str, cell_style_centered),
                Paragraph(lunch_break_str, cell_style_centered),
                Paragraph(qr_info, cell_style),
                Paragraph(' '.join(notes) if notes else '✓', notes_style)
            ])
        
        current_date += timedelta(days=1)
    
    # Create table - optimized column widths (A4 landscape is 29.7cm, minus 2cm margins = 27.7cm)
    table = Table(table_data, colWidths=[2.5*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.7*cm, 7.3*cm, 7.6*cm])
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('LEFTPADDING', (0, 0), (-1, 0), 6),
        ('RIGHTPADDING', (0, 0), (-1, 0), 6),
        
        # Data rows styling
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 1), (-1, -1), 6),
        ('RIGHTPADDING', (0, 1), (-1, -1), 6),
        
        # Alignment for specific columns
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Date
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Day
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Arrival
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Departure
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Hours
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),    # Location
        ('ALIGN', (6, 1), (6, -1), 'LEFT'),    # Notes
        
        # Grid and backgrounds
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1e40af')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        
        # Border styling
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6b7280')),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    summary_elements = []
    summary_elements.append(Paragraph(str(_('Summary Statistics')), heading_style))
    
    avg_hours = total_work_hours / total_days if total_days > 0 else 0
    
    # Create styled summary label and value styles
    summary_label_style = ParagraphStyle(
        'SummaryLabel',
        parent=styles['Normal'],
        fontName=font_name_bold,
        fontSize=10,
        leading=13
    )
    
    summary_value_style = ParagraphStyle(
        'SummaryValue',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=10,
        leading=13
    )
    
    # Calculate overtime (based on monthly work_hours)
    # Calculate expected hours based on the date range
    months_in_range = set()
    temp_date = date_from.date()
    while temp_date <= date_to.date():
        months_in_range.add((temp_date.year, temp_date.month))
        temp_date += timedelta(days=1)
    
    # Calculate expected hours proportionally
    expected_hours = 0
    for year, month in months_in_range:
        days_in_month = monthrange(year, month)[1]
        # Days in this month that are in our range
        month_start = max(date_from.date(), datetime(year, month, 1).date())
        month_end = min(date_to.date(), datetime(year, month, days_in_month).date())
        days_in_range = (month_end - month_start).days + 1
        
        # Calculate proportional expected hours
        month_expected = (user.working_hours / days_in_month) * days_in_range
        expected_hours += month_expected
    
    overtime_hours = max(0, total_work_hours - expected_hours)
    
    # Create styled summary data
    summary_data = [
        [
            Paragraph(str(_('Total Working Days')), summary_label_style),
            Paragraph(str(total_days), summary_value_style)
        ],
        [
            Paragraph(str(_('Expected Hours')), summary_label_style),
            Paragraph(f"{int(expected_hours)}:{int((expected_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Total Hours (with breaks)')), summary_label_style),
            Paragraph(f"{int(total_work_hours_with_breaks)}:{int((total_work_hours_with_breaks % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Total Hours (without breaks)')), summary_label_style),
            Paragraph(f"{int(total_work_hours)}:{int((total_work_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Total Break Time')), summary_label_style),
            Paragraph(f"{int(total_break_minutes // 60)}:{int(total_break_minutes % 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Overtime Hours')), summary_label_style),
            Paragraph(f"{int(overtime_hours)}:{int((overtime_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Night Hours (22:00-06:00)')), summary_label_style),
            Paragraph(f"{int(total_night_hours)}:{int((total_night_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Holiday Hours')), summary_label_style),
            Paragraph(f"{int(total_holiday_hours)}:{int((total_holiday_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Average Hours per Day')), summary_label_style),
            Paragraph(f"{int(avg_hours)}:{int((avg_hours % 1) * 60):02d}", summary_value_style)
        ],
        [
            Paragraph(str(_('Vacation Days')), summary_label_style),
            Paragraph(str(total_vacation_days), summary_value_style)
        ],
        [
            Paragraph(str(_('Home Office Days')), summary_label_style),
            Paragraph(str(total_home_office_days), summary_value_style)
        ],
        [
            Paragraph(str(_('Days with Issues')), summary_label_style),
            Paragraph(str(len(set(days_with_issues))), summary_value_style)
        ],
    ]
    
    summary_table = Table(summary_data, colWidths=[10*cm, 7*cm])
    summary_table.setStyle(TableStyle([
        # Background colors
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        
        # Font styling
        ('FONTNAME', (0, 0), (0, -1), font_name_bold),
        ('FONTNAME', (1, 0), (1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        
        # Alignment
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#6b7280')),
        ('LINEAFTER', (0, 0), (0, -1), 1, colors.HexColor('#a5b4fc')),
    ]))
    
    summary_elements.append(summary_table)
    
    # Add summary as a single block that won't be split across pages
    elements.append(KeepTogether(summary_elements))
    
    # Build PDF
    doc.build(elements)
    
    # Return PDF response - serve from file
    with open(filepath, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    return response


def generate_attendance_excel(request, user_id):
    """Generate Excel attendance report for a user"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, timedelta
    from collections import defaultdict
    from calendar import monthrange
    import holidays
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        current_user = crud.get_user_by_id(request.session['user_id'])
        if not current_user or not current_user.is_manager or not current_user.can_edit_employees:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = current_user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    user = crud.get_user_by_id(user_id, include_inactive=True)
    if not user or user.company != company:
        messages.error(request, _('User not found'))
        return redirect('company_dashboard')
    
    # Parse date range - accept both date_range and date_from/date_to parameters
    date_range = request.GET.get('date_range', '')
    date_from_param = request.GET.get('date_from', '')
    date_to_param = request.GET.get('date_to', '')
    
    try:
        if date_from_param and date_to_param:
            # Use date_from and date_to parameters (format: YYYY-MM-DD)
            date_from = datetime.strptime(date_from_param, '%Y-%m-%d')
            date_to = datetime.strptime(date_to_param, '%Y-%m-%d')
        elif date_range and ' - ' in date_range:
            # Use date_range parameter (format: DD.MM.YYYY - DD.MM.YYYY)
            date_from_str, date_to_str = date_range.split(' - ')
            date_from = datetime.strptime(date_from_str.strip(), '%d.%m.%Y')
            date_to = datetime.strptime(date_to_str.strip(), '%d.%m.%Y')
        else:
            messages.error(request, _('Invalid date range'))
            return redirect('view_user_details', user_id=user_id)
    except:
        messages.error(request, _('Invalid date format'))
        return redirect('view_user_details', user_id=user_id)
    
    # Get scans in date range
    scans = ScanEvent.objects.filter(
        scanned_by=user,
        timestamp__date__gte=date_from.date(),
        timestamp__date__lte=date_to.date()
    ).select_related('qr_code').order_by('timestamp')
    
    vacations = Vacation.objects.filter(
        user=user,
        is_active=True,
        date_from__lte=date_to.date(),
        date_to__gte=date_from.date()
    ).order_by('date_from')
    
    # Create directory structure for Excel storage
    now = datetime.now()
    excel_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(excel_dir, exist_ok=True)
    
    # Generate filename
    filename = f"attendance_{user.name.replace(' ', '_')}_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(excel_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = str(_('Attendance Report'))
    
    # Define styles
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    title_font = Font(color='2563eb', bold=True, size=16)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='d1d5db'),
        right=Side(style='thin', color='d1d5db'),
        top=Side(style='thin', color='d1d5db'),
        bottom=Side(style='thin', color='d1d5db')
    )
    
    # Title and header information
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{_('Attendance Report')} - {user.name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"{_('Period')}: {date_from.strftime('%d.%m.%Y')} - {date_to.strftime('%d.%m.%Y')}"
    ws['A2'].alignment = center_align
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"{_('Company')}: {company.name}"
    ws['A3'].alignment = center_align
    
    # Table headers (row 5)
    headers = [
        str(_('Date')),
        str(_('Day')),
        str(_('Arrival')),
        str(_('Departure')),
        str(_('Hours')),
        str(_('Break')),
        str(_('Scanned QR')),
        str(_('Notes'))
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Group scans by day
    daily_data = defaultdict(list)
    for scan in scans:
        day = scan.timestamp.date()
        daily_data[day].append(scan)
    
    # Create dictionary of vacation days with vacation object
    vacation_days = {}
    for vacation in vacations:
        current = vacation.date_from
        while current <= vacation.date_to:
            vacation_days[current] = vacation
            current += timedelta(days=1)
    
    # Helper function to calculate night hours (22:00-06:00)
    def calculate_night_hours(start_time, end_time):
        """Calculate hours worked between 22:00 and 06:00"""
        night_hours = 0
        current = start_time
        
        while current < end_time:
            # Check if current hour is night time (22:00-23:59 or 00:00-05:59)
            hour = current.hour
            if hour >= 22 or hour < 6:
                # Calculate minutes in this hour that count as night work
                next_hour = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
                segment_end = min(next_hour, end_time)
                segment_duration = (segment_end - current).total_seconds() / 3600
                night_hours += segment_duration
            
            # Move to next hour
            current = current.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            if current >= end_time:
                break
        
        return night_hours
    
    # Calculate statistics
    total_days = len(daily_data)
    total_work_hours = 0
    total_work_hours_with_breaks = 0
    total_night_hours = 0
    total_break_minutes = 0
    days_with_issues = []
    total_vacation_days = 0
    total_home_office_days = 0
    total_holiday_hours = 0
    
    # Get holidays based on language
    lang_code = request.LANGUAGE_CODE if hasattr(request, 'LANGUAGE_CODE') else 'sk'
    if lang_code == 'sk':
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    elif lang_code == 'en':
        country_holidays = holidays.UnitedStates(years=[date_from.year, date_to.year])
    elif lang_code == 'de':
        country_holidays = holidays.Germany(years=[date_from.year, date_to.year])
    elif lang_code == 'es':
        country_holidays = holidays.Spain(years=[date_from.year, date_to.year])
    else:
        country_holidays = holidays.Slovakia(years=[date_from.year, date_to.year])
    
    # Populate daily attendance data
    row = 6
    current_date = date_from.date()
    alt_fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')
    vacation_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
    sick_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
    doctor_fill = PatternFill(start_color='ede9fe', end_color='ede9fe', fill_type='solid')
    warning_font = Font(color='f59e0b')
    
    while current_date <= date_to.date():
        day_scans = daily_data.get(current_date, [])
        day_name = date_format(current_date, format='l')
        
        # Check if this day is a holiday
        is_holiday = current_date in country_holidays
        holiday_name = country_holidays.get(current_date, '') if is_holiday else ''
        
        # Check if this day is a vacation day
        vacation_obj = vacation_days.get(current_date)
        is_vacation = vacation_obj is not None
        vacation_type = vacation_obj.type if vacation_obj else None
        if is_vacation:
            if vacation_type == 'home_office':
                total_home_office_days += 1
            else:
                total_vacation_days += 1
        
        # Date
        ws.cell(row=row, column=1, value=current_date.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        # Day
        ws.cell(row=row, column=2, value=day_name)
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        
        if is_vacation and not day_scans:
            # Vacation day with no scans
            ws.cell(row=row, column=3, value='-')
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value='-')
            
            # Format time if available
            time_info = ""
            if vacation_obj and vacation_obj.time_from and vacation_obj.time_to:
                time_info = f" ({vacation_obj.time_from.strftime('%H:%M')} - {vacation_obj.time_to.strftime('%H:%M')})"
            
            if vacation_type == 'sick_leave':
                ws.cell(row=row, column=8, value=f"🏥 {_('Sick Leave')}{time_info}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = sick_fill
            elif vacation_type == 'doctor':
                ws.cell(row=row, column=8, value=f"👨‍⚕️ {_('Doctor')}{time_info}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = doctor_fill
            elif vacation_type == 'home_office':
                ws.cell(row=row, column=8, value=f"🏠 {_('Home Office')}{time_info}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
            else:
                ws.cell(row=row, column=8, value=f"🏖 {_('Vacation')}{time_info}")
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = vacation_fill
        
        elif not day_scans and not is_vacation:
            # No scans for this day
            ws.cell(row=row, column=3, value='-')
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='0:00')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value='-')
            ws.cell(row=row, column=8, value=str(_('No scans')))
        
        else:
            # Day with scans
            arrivals = [s for s in day_scans if s.scan_type == 'arrival']
            departures = [s for s in day_scans if s.scan_type == 'departure']
            
            # Check for issues
            notes = []
            
            if is_vacation:
                if vacation_type == 'sick_leave':
                    notes.append(f"⚠ {_('Scans on sick leave day')}")
                elif vacation_type == 'doctor':
                    notes.append(f"⚠ {_('Scans on doctor day')}")
                elif vacation_type == 'home_office':
                    notes.append(f"🏠 {_('Home Office')}")
                else:
                    notes.append(f"⚠ {_('Scans on vacation day')}")
                # Only mark as issue if not home office
                if vacation_type != 'home_office':
                    days_with_issues.append(current_date)
            
            # Add holiday note
            if is_holiday:
                notes.append(f"🎉 {holiday_name}")
            
            if not arrivals:
                notes.append(f"⚠ {_('Missing arrival')}")
                days_with_issues.append(current_date)
            if not departures:
                notes.append(f"⚠ {_('Missing departure')}")
                days_with_issues.append(current_date)
            
            # Calculate hours worked and lunch breaks
            hours_worked = 0
            night_hours = 0
            lunch_break_minutes = 0
            
            if arrivals and departures:
                first_arrival = arrivals[0].timestamp
                last_departure = departures[-1].timestamp
                work_duration = last_departure - first_arrival
                hours_worked = work_duration.total_seconds() / 3600
                night_hours = calculate_night_hours(first_arrival, last_departure)
                
                # Calculate lunch break
                # First check if there are actual lunch break scans
                lunch_starts = [s for s in day_scans if s.scan_type == 'lunch_break_start']
                lunch_ends = [s for s in day_scans if s.scan_type == 'lunch_break_end']
                
                if lunch_starts and lunch_ends:
                    # Use actual scanned lunch breaks
                    for i in range(min(len(lunch_starts), len(lunch_ends))):
                        break_duration = lunch_ends[i].timestamp - lunch_starts[i].timestamp
                        lunch_break_minutes += break_duration.total_seconds() / 60
                elif company.auto_lunch_breaks and user.has_lunch_break:
                    # No scanned breaks, use default duration if auto breaks enabled
                    lunch_break_minutes = user.lunch_break_duration
                # else: lunch_break_minutes stays 0
                
                # Track hours with and without breaks
                hours_without_break = hours_worked - (lunch_break_minutes / 60)
                total_work_hours += hours_without_break  # Hours without break
                total_work_hours_with_breaks += hours_worked  # Hours including break
                total_night_hours += night_hours
                
                # Track holiday hours
                if is_holiday:
                    total_holiday_hours += hours_worked
            
            # Format times
            arrival_time = arrivals[0].timestamp.strftime('%H:%M') if arrivals else '-'
            departure_time = departures[-1].timestamp.strftime('%H:%M') if departures else '-'
            hours_str = f"{int(hours_worked)}:{int((hours_worked % 1) * 60):02d}" if hours_worked > 0 else '0:00'
            # Format break time as HH:MM
            if lunch_break_minutes > 0:
                break_hours = int(lunch_break_minutes // 60)
                break_mins = int(lunch_break_minutes % 60)
                lunch_break_str = f"{break_hours}:{break_mins:02d}"
                total_break_minutes += lunch_break_minutes
            else:
                lunch_break_str = '-'
            
            # Get QR code info
            if arrivals:
                if arrivals[0].is_home_office:
                    qr_info = f"{_('Home Office')} - {_('Home Office')}"
                elif arrivals[0].is_business_trip:
                    qr_info = f"{_('Business Trip')} - {_('Business Trip')}"
                else:
                    qr_info = f"{arrivals[0].qr_code.name} - {arrivals[0].qr_code.location}"
            elif departures:
                if departures[0].is_home_office:
                    qr_info = f"{_('Home Office')} - {_('Home Office')}"
                elif departures[0].is_business_trip:
                    qr_info = f"{_('Business Trip')} - {_('Business Trip')}"
                else:
                    qr_info = f"{departures[0].qr_code.name} - {departures[0].qr_code.location}"
            else:
                qr_info = '-'
            
            ws.cell(row=row, column=3, value=arrival_time)
            ws.cell(row=row, column=4, value=departure_time)
            ws.cell(row=row, column=5, value=hours_str)
            ws.cell(row=row, column=6, value=lunch_break_str)
            ws.cell(row=row, column=7, value=qr_info)
            ws.cell(row=row, column=8, value=' '.join(notes) if notes else '✓')
            
            if notes:
                ws.cell(row=row, column=8).font = warning_font
        
        # Apply alignment and borders
        for col in range(3, 9):
            if col in [3, 4, 5, 6]:  # Center align time and break columns
                ws.cell(row=row, column=col).alignment = center_align
            else:
                ws.cell(row=row, column=col).alignment = left_align
            ws.cell(row=row, column=col).border = thin_border
        
        # Alternate row colors
        if row % 2 == 0 and not is_vacation:
            for col in range(1, 9):
                if ws.cell(row=row, column=col).fill.start_color.rgb != 'd1fae5' and \
                   ws.cell(row=row, column=col).fill.start_color.rgb != 'fef3c7':
                    ws.cell(row=row, column=col).fill = alt_fill
        
        current_date += timedelta(days=1)
        row += 1
    
    # Add summary section
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row=row, column=1, value=str(_('Summary Statistics')))
    ws.cell(row=row, column=1).font = Font(color='1e40af', bold=True, size=14)
    ws.cell(row=row, column=1).alignment = left_align
    
    row += 1
    summary_fill = PatternFill(start_color='e0e7ff', end_color='e0e7ff', fill_type='solid')
    
    avg_hours = total_work_hours / total_days if total_days > 0 else 0
    
    # Calculate overtime (based on monthly work_hours)
    months_in_range = set()
    temp_date = date_from.date()
    while temp_date <= date_to.date():
        months_in_range.add((temp_date.year, temp_date.month))
        temp_date += timedelta(days=1)
    
    # Calculate expected hours proportionally
    expected_hours = 0
    for year, month in months_in_range:
        days_in_month = monthrange(year, month)[1]
        month_start = max(date_from.date(), datetime(year, month, 1).date())
        month_end = min(date_to.date(), datetime(year, month, days_in_month).date())
        days_in_range = (month_end - month_start).days + 1
        month_expected = (user.working_hours / days_in_month) * days_in_range
        expected_hours += month_expected
    
    overtime_hours = max(0, total_work_hours - expected_hours)
    
    summary_data = [
        (str(_('Total Working Days')), str(total_days)),
        (str(_('Expected Hours')), f"{int(expected_hours)}:{int((expected_hours % 1) * 60):02d}"),
        (str(_('Total Hours (with breaks)')), f"{int(total_work_hours_with_breaks)}:{int((total_work_hours_with_breaks % 1) * 60):02d}"),
        (str(_('Total Hours (without breaks)')), f"{int(total_work_hours)}:{int((total_work_hours % 1) * 60):02d}"),
        (str(_('Total Break Time')), f"{int(total_break_minutes // 60)}:{int(total_break_minutes % 60):02d}"),
        (str(_('Overtime Hours')), f"{int(overtime_hours)}:{int((overtime_hours % 1) * 60):02d}"),
        (str(_('Night Hours (22:00-06:00)')), f"{int(total_night_hours)}:{int((total_night_hours % 1) * 60):02d}"),
        (str(_('Holiday Hours')), f"{int(total_holiday_hours)}:{int((total_holiday_hours % 1) * 60):02d}"),
        (str(_('Average Hours per Day')), f"{int(avg_hours)}:{int((avg_hours % 1) * 60):02d}"),
        (str(_('Vacation Days')), str(total_vacation_days)),
        (str(_('Home Office Days')), str(total_home_office_days)),
        (str(_('Days with Issues')), str(len(set(days_with_issues)))),
    ]
    
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=1).fill = summary_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = left_align
        
        ws.merge_cells(f'B{row}:H{row}')
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = center_align
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 25
    
    # Save workbook
    wb.save(filepath)
    
    # Return Excel response
    with open(filepath, 'rb') as excel_file:
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def generate_qr_code_pdf(request, qr_id):
    """Generate PDF with QR code for printing on A4"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from datetime import datetime
    import os
    
    is_company = request.session.get('user_type') == 'company' and 'company_id' in request.session
    is_manager = request.session.get('user_type') == 'user' and 'user_id' in request.session
    
    if not (is_company or is_manager):
        messages.error(request, _('Unauthorized'))
        return redirect('company_login')
    
    # Get company
    if is_company:
        company = crud.get_company_by_id(request.session['company_id'])
    else:
        user = crud.get_user_by_id(request.session['user_id'])
        if not user or not user.is_manager or not user.can_edit_qr_codes:
            messages.error(request, _('Access denied'))
            return redirect('user_dashboard')
        company = user.company
    
    if not company:
        messages.error(request, _('Company not found'))
        return redirect('company_login' if is_company else 'user_login')
    
    qr_code = crud.get_qr_code_by_id(qr_id)
    if not qr_code or qr_code.company != company:
        messages.error(request, _('QR Code not found'))
        return redirect('company_dashboard')
    
    # Create directory structure for PDF storage
    now = datetime.now()
    pdf_dir = os.path.join(settings.MEDIA_ROOT, 'PDF', str(now.year), f"{now.month:02d}")
    os.makedirs(pdf_dir, exist_ok=True)
    
    # Generate filename
    filename = f"qr_code_{qr_code.name.replace(' ', '_')}_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(pdf_dir, filename)
    
    # Register DejaVu fonts for Unicode support
    try:
        font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
        dejavu_path = os.path.join(font_dir, 'DejaVuSans.ttf')
        dejavu_bold_path = os.path.join(font_dir, 'DejaVuSans-Bold.ttf')
        
        if os.path.exists(dejavu_path) and os.path.exists(dejavu_bold_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans', dejavu_path))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', dejavu_bold_path))
            font_name_bold = 'DejaVuSans-Bold'
        else:
            raise Exception("DejaVu fonts not found")
    except:
        font_name_bold = 'Helvetica-Bold'
    
    # Create PDF
    doc = SimpleDocTemplate(filepath, pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=3*cm, bottomMargin=3*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style - centered
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=24,
        alignment=TA_CENTER,
        spaceAfter=30,
    )
    
    # Add title
    title = Paragraph(qr_code.name, title_style)
    elements.append(title)
    
    # Add spacer
    elements.append(Spacer(1, 2*cm))
    
    # Add QR code image - centered and larger
    qr_image_path = os.path.join(settings.MEDIA_ROOT, qr_code.qr_code.name)
    if os.path.exists(qr_image_path):
        # Create image with specific size (12cm x 12cm)
        img = Image(qr_image_path, width=12*cm, height=12*cm)
        img.hAlign = 'CENTER'
        elements.append(img)
    
    # Build PDF
    doc.build(elements)
    
    # Return PDF response - open in new tab
    with open(filepath, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    return response
